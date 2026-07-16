from __future__ import annotations

import hashlib
import io
import json
import re
import secrets
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app import config
from app.services.package_service import PackageInput, create_package
from app.services.recipe_import_service import SECTION_ORDER
from app.utils import database


EDITABLE_RECIPE_STATUSES = {"IMPORTED", "NEEDS_REVIEW"}
IMMUTABLE_RECIPE_STATUSES = {"APPROVED", "ACTIVE", "SUPERSEDED", "ARCHIVED"}
APPROVED_ASSIGNMENTS = {"APPROVED", "REPLACED"}
VALID_INTEGRITY_STATUSES = {config.DOCUMENT_STATUS_DOWNLOADED, "VERIFIED"}


def _token(prefix: str) -> str:
    return f"{prefix}-{secrets.token_hex(8).upper()}"


def _json(value: Any) -> str:
    return json.dumps(value, sort_keys=True, default=str)


def _loads(value: Any, default: Any) -> Any:
    try:
        return json.loads(value) if value else default
    except (TypeError, json.JSONDecodeError):
        return default


def _audit(
    event_type: str,
    *,
    actor: str,
    db_path: Path | str,
    package_id: str | None = None,
    recipe_id: str | None = None,
    slot_instance_id: str | None = None,
    document_id: str | None = None,
    details: dict[str, Any] | None = None,
) -> None:
    safe_details = {key: value for key, value in (details or {}).items() if key.lower() not in {"content", "text", "api_key", "authorization"}}
    with database.get_connection(db_path) as connection:
        connection.execute(
            "INSERT INTO phase6a_audit_events VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (_token("P6A"), package_id, recipe_id, slot_instance_id, document_id, event_type,
             _json(safe_details), actor, database.utc_now_iso()),
        )


def _decoded_recipe(row: dict[str, Any]) -> dict[str, Any]:
    return dict(row)


def _decoded_slot(row: dict[str, Any]) -> dict[str, Any]:
    slot = dict(row)
    for source, target in (
        ("preferred_sources_json", "preferred_sources"),
        ("fallback_sources_json", "fallback_sources"),
        ("allowed_document_types_json", "allowed_document_types"),
        ("source_coordinates_json", "source_coordinates"),
        ("raw_import_json", "raw_import"),
    ):
        slot[target] = _loads(slot.get(source), [] if "sources" in source or "types" in source else {})
    return slot


def get_recipe(recipe_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any] | None:
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        row = connection.execute("SELECT * FROM package_recipes WHERE recipe_id = ?", (recipe_id,)).fetchone()
    return _decoded_recipe(dict(row)) if row else None


def list_recipes(*, db_path: Path | str = config.DATABASE_PATH) -> list[dict[str, Any]]:
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        rows = connection.execute("SELECT * FROM package_recipes ORDER BY recipe_name, version DESC").fetchall()
    return [_decoded_recipe(dict(row)) for row in rows]


def list_recipe_slots(recipe_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> list[dict[str, Any]]:
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        rows = connection.execute(
            "SELECT * FROM research_slots WHERE recipe_id = ? ORDER BY order_number, suborder, source_row",
            (recipe_id,),
        ).fetchall()
    return [_decoded_slot(dict(row)) for row in rows]


def get_active_recipe(*, security_type: str = "Common Equity", db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any] | None:
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        row = connection.execute(
            "SELECT * FROM package_recipes WHERE security_type = ? AND status = 'ACTIVE' ORDER BY version DESC LIMIT 1",
            (security_type,),
        ).fetchone()
    return dict(row) if row else None


def update_draft_slot(
    slot_id: str,
    updates: dict[str, Any],
    *,
    actor: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    database.initialize_database(db_path)
    allowed = {
        "display_name", "section_code", "section_name", "required_level", "long_applicable", "short_applicable",
        "conditional_rule", "preferred_sources_json", "fallback_sources_json", "instructions", "minimum_documents",
        "maximum_documents", "freshness_rule", "anchor_rule", "allowed_document_types_json", "expected_output_format",
        "auto_search_enabled", "manual_upload_allowed", "analyst_review_required", "default_status", "enabled", "import_warning",
    }
    selected = {key: value for key, value in updates.items() if key in allowed}
    with database.get_connection(db_path) as connection:
        row = connection.execute(
            "SELECT s.*, r.status AS recipe_status FROM research_slots s JOIN package_recipes r ON r.recipe_id=s.recipe_id WHERE s.slot_id=?",
            (slot_id,),
        ).fetchone()
        if not row:
            raise ValueError("Recipe slot does not exist.")
        if row["recipe_status"] not in EDITABLE_RECIPE_STATUSES:
            raise ValueError("Approved and active recipes are immutable. Create a new draft version.")
        if selected:
            sql = ", ".join(f"{key} = ?" for key in selected)
            connection.execute(f"UPDATE research_slots SET {sql} WHERE slot_id = ?", (*selected.values(), slot_id))
        refreshed = connection.execute("SELECT * FROM research_slots WHERE slot_id = ?", (slot_id,)).fetchone()
    return _decoded_slot(dict(refreshed))


def create_draft_version(
    recipe_id: str,
    *,
    created_by: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    recipe = get_recipe(recipe_id, db_path=db_path)
    if not recipe:
        raise ValueError("Recipe does not exist.")
    slots = list_recipe_slots(recipe_id, db_path=db_path)
    new_id = _token("RCP")
    now = database.utc_now_iso()
    with database.get_connection(db_path) as connection:
        version = int(connection.execute(
            "SELECT COALESCE(MAX(version), 0) + 1 FROM package_recipes WHERE recipe_name = ?",
            (recipe["recipe_name"],),
        ).fetchone()[0])
        fields = (
            "recipe_id", "recipe_name", "recipe_type", "security_type", "version", "description", "source_workbook_name",
            "source_workbook_hash", "source_sheet", "importer_version", "status", "created_at", "created_by", "notes", "import_id",
        )
        values = (new_id, recipe["recipe_name"], recipe["recipe_type"], recipe["security_type"], version, recipe.get("description"),
                  recipe["source_workbook_name"], recipe["source_workbook_hash"], recipe["source_sheet"], recipe["importer_version"],
                  "NEEDS_REVIEW", now, created_by, f"Drafted from immutable recipe {recipe_id}.", recipe.get("import_id"))
        connection.execute(f"INSERT INTO package_recipes({', '.join(fields)}) VALUES ({', '.join('?' for _ in fields)})", values)
        for slot in slots:
            columns = [column for column in slot if column not in {"slot_id", "recipe_id", "preferred_sources", "fallback_sources", "allowed_document_types", "source_coordinates", "raw_import"}]
            copied = {column: slot[column] for column in columns}
            copied.update({"slot_id": _token("SLOT"), "recipe_id": new_id, "created_at": now})
            insert_columns = list(copied)
            connection.execute(
                f"INSERT INTO research_slots({', '.join(insert_columns)}) VALUES ({', '.join('?' for _ in insert_columns)})",
                tuple(copied[column] for column in insert_columns),
            )
    return get_recipe(new_id, db_path=db_path) or {}


def approve_recipe(recipe_id: str, *, approver: str, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    recipe = get_recipe(recipe_id, db_path=db_path)
    if not recipe or recipe["status"] not in EDITABLE_RECIPE_STATUSES:
        raise ValueError("Only a reviewable draft recipe can be approved.")
    slots = list_recipe_slots(recipe_id, db_path=db_path)
    if not slots or any(not slot.get("enabled") for slot in slots if slot.get("required_level") == "REQUIRED"):
        raise ValueError("The recipe must contain enabled slots before approval.")
    now = database.utc_now_iso()
    snapshot = {"recipe": recipe, "slots": slots}
    with database.get_connection(db_path) as connection:
        connection.execute(
            "UPDATE package_recipes SET status='APPROVED', approved_at=?, approved_by=? WHERE recipe_id=?",
            (now, approver, recipe_id),
        )
        connection.execute(
            "INSERT INTO recipe_approvals VALUES (?, ?, ?, ?, ?, ?, ?)",
            (_token("RAPP"), recipe_id, approver, now, recipe["source_workbook_hash"], recipe["version"], _json(snapshot)),
        )
    _audit("RECIPE_APPROVED", actor=approver, recipe_id=recipe_id, details={"version": recipe["version"], "workbook_sha256": recipe["source_workbook_hash"]}, db_path=db_path)
    return get_recipe(recipe_id, db_path=db_path) or {}


def activate_recipe(recipe_id: str, *, actor: str, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    recipe = get_recipe(recipe_id, db_path=db_path)
    if not recipe or recipe["status"] not in {"APPROVED", "ACTIVE"}:
        raise ValueError("Approve the recipe before activation.")
    now = database.utc_now_iso()
    with database.get_connection(db_path) as connection:
        active = connection.execute(
            "SELECT recipe_id FROM package_recipes WHERE security_type=? AND status='ACTIVE' AND recipe_id != ?",
            (recipe["security_type"], recipe_id),
        ).fetchall()
        for row in active:
            connection.execute(
                "UPDATE package_recipes SET status='SUPERSEDED', superseded_at=?, superseded_by_recipe_id=? WHERE recipe_id=?",
                (now, recipe_id, row["recipe_id"]),
            )
        connection.execute("UPDATE package_recipes SET status='ACTIVE' WHERE recipe_id=?", (recipe_id,))
    for row in active:
        _audit("RECIPE_SUPERSEDED", actor=actor, recipe_id=row["recipe_id"], details={"superseded_by": recipe_id}, db_path=db_path)
    _audit("RECIPE_ACTIVATED", actor=actor, recipe_id=recipe_id, details={"version": recipe["version"]}, db_path=db_path)
    return get_recipe(recipe_id, db_path=db_path) or {}


def instantiate_recipe(
    package_id: str,
    recipe_id: str,
    *,
    created_by: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    recipe = get_recipe(recipe_id, db_path=db_path)
    if not recipe or recipe["status"] not in {"APPROVED", "ACTIVE", "SUPERSEDED"}:
        raise ValueError("Only an approved recipe can be instantiated.")
    slots = [slot for slot in list_recipe_slots(recipe_id, db_path=db_path) if slot.get("enabled")]
    instance_id = _token("PRI")
    now = database.utc_now_iso()
    snapshot = {"recipe": recipe, "slots": slots}
    with database.get_connection(db_path) as connection:
        connection.execute(
            "INSERT INTO package_recipe_instances VALUES (?, ?, ?, ?, ?, ?, ?, 'DRAFT', NULL, NULL)",
            (instance_id, package_id, recipe_id, recipe["version"], _json(snapshot), now, created_by),
        )
        for slot in slots:
            applicability = "NEEDS_REVIEW" if slot["required_level"] == "CONDITIONAL" else "APPLICABLE"
            completion = "NEEDS_ANALYST_REVIEW" if applicability == "NEEDS_REVIEW" else slot["default_status"]
            connection.execute(
                """INSERT INTO package_slot_instances(
                    package_slot_instance_id, package_recipe_instance_id, package_id, slot_id, order_number, suborder,
                    display_name_snapshot, section_snapshot, requirement_snapshot, instructions_snapshot,
                    preferred_sources_snapshot_json, minimum_documents, maximum_documents, applicability_status,
                    completion_status, analyst_acknowledged, analyst_notes, selected_document_count,
                    latest_selected_document_date, cap_override_approved, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, NULL, 0, NULL, 0, ?, ?)""",
                (_token("PSI"), instance_id, package_id, slot["slot_id"], slot["order_number"], slot["suborder"],
                 slot["display_name"], slot["section_name"], slot["required_level"], slot.get("instructions"),
                 slot["preferred_sources_json"], slot["minimum_documents"], slot["maximum_documents"],
                 applicability, completion, now, now),
            )
    _audit("PACKAGE_RECIPE_INSTANTIATED", actor=created_by, package_id=package_id, recipe_id=recipe_id,
           details={"recipe_version": recipe["version"], "slot_count": len(slots)}, db_path=db_path)
    return get_package_recipe_instance(package_id, db_path=db_path) or {}


def create_package_from_active_recipe(
    company_metadata: dict[str, Any],
    *,
    research_cutoff: date,
    compilation_date: date,
    compiled_by: str,
    created_by: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    recipe = get_active_recipe(db_path=db_path)
    if not recipe:
        raise ValueError("No active Common Equity recipe is available. An administrator must approve and activate one.")
    package = create_package(
        PackageInput(
            ticker=str(company_metadata.get("ticker") or ""), security_type="Common Equity",
            research_cutoff_date=research_cutoff, filing_history_years=3, analyst_notes="",
        ),
        db_path=db_path,
    )
    package = database.update_package_company_metadata(package["package_id"], company_metadata, db_path=db_path) or package
    with database.get_connection(db_path) as connection:
        connection.execute(
            "UPDATE packages SET compilation_date=?, compiled_by=?, updated_at=? WHERE package_id=?",
            (compilation_date.isoformat(), compiled_by.strip(), database.utc_now_iso(), package["package_id"]),
        )
    instantiate_recipe(package["package_id"], recipe["recipe_id"], created_by=created_by, db_path=db_path)
    return database.get_package_by_package_id(package["package_id"], db_path=db_path) or package


def get_package_recipe_instance(package_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any] | None:
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        row = connection.execute("SELECT * FROM package_recipe_instances WHERE package_id=?", (package_id,)).fetchone()
    if not row:
        return None
    result = dict(row)
    result["recipe_snapshot"] = _loads(result.get("recipe_snapshot_json"), {})
    return result


def list_slot_instances(package_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> list[dict[str, Any]]:
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        rows = connection.execute(
            """SELECT psi.*, rs.section_code, rs.normalized_slot_type, rs.manual_upload_allowed,
                      rs.auto_search_enabled, rs.freshness_rule, rs.anchor_rule,
                      rs.allowed_document_types_json, rs.analyst_review_required
               FROM package_slot_instances psi JOIN research_slots rs ON rs.slot_id=psi.slot_id
               WHERE psi.package_id=? ORDER BY psi.order_number, psi.suborder, psi.created_at""",
            (package_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def list_assignments(package_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> list[dict[str, Any]]:
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        rows = connection.execute(
            """SELECT a.*, psi.order_number AS slot_order_number, psi.suborder AS slot_suborder,
                      d.title AS document_title, d.document_date, d.publication_date, d.source_name,
                      d.local_path, d.local_filename, d.original_filename, d.mime_type, d.file_size_bytes,
                      d.sha256_hash, d.source_url, d.collection_status
               FROM slot_document_assignments a
               JOIN package_slot_instances psi ON psi.package_slot_instance_id=a.package_slot_instance_id
               JOIN documents d ON d.document_id=a.document_id
               WHERE a.package_id=? ORDER BY psi.order_number, psi.suborder, a.display_order, a.assigned_at""",
            (package_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def _freshness_violation(slot: dict[str, Any], assignments: list[dict[str, Any]]) -> bool:
    rule = str(slot.get("freshness_rule") or "")
    match = re.search(r"max_age_days\s*[:=]\s*(\d+)", rule, re.I)
    if not match:
        return False
    limit = int(match.group(1))
    for assignment in assignments:
        raw_date = assignment.get("document_date") or assignment.get("publication_date")
        try:
            if raw_date and (date.today() - date.fromisoformat(str(raw_date)[:10])).days > limit:
                return True
        except ValueError:
            continue
    return False


def recalculate_completion(package_id: str, *, actor: str = "system", db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    slots = list_slot_instances(package_id, db_path=db_path)
    assignments = list_assignments(package_id, db_path=db_path)
    by_slot: dict[str, list[dict[str, Any]]] = {}
    for assignment in assignments:
        if assignment["assignment_status"] not in {"REJECTED", "REPLACED", "REMOVED"}:
            by_slot.setdefault(assignment["package_slot_instance_id"], []).append(assignment)
    now = database.utc_now_iso()
    updates: list[tuple[Any, ...]] = []
    for slot in slots:
        selected = [item for item in by_slot.get(slot["package_slot_instance_id"], []) if item.get("selected_for_package")]
        approved = [item for item in selected if item["assignment_status"] == "APPROVED" and item["collection_status"] in VALID_INTEGRITY_STATUSES]
        count = len(approved)
        dates = [str(item.get("document_date") or item.get("publication_date")) for item in approved if item.get("document_date") or item.get("publication_date")]
        if slot["applicability_status"] == "NOT_APPLICABLE":
            status = "NOT_APPLICABLE"
        elif slot["analyst_acknowledged"] and slot["completion_status"] == "NOT_AVAILABLE":
            status = "NOT_AVAILABLE"
        elif slot["applicability_status"] == "NEEDS_REVIEW":
            status = "NEEDS_ANALYST_REVIEW"
        elif _freshness_violation(slot, approved):
            status = "STALE"
        elif count > slot["maximum_documents"] and not slot["cap_override_approved"]:
            status = "NEEDS_ANALYST_REVIEW"
        elif count >= slot["minimum_documents"] and all(item["assignment_status"] == "APPROVED" for item in selected):
            status = "COMPLETE"
        elif selected:
            status = "PARTIAL"
        elif slot["requirement_snapshot"] == "REQUIRED":
            status = "MISSING"
        elif slot["manual_upload_allowed"]:
            status = "MANUAL_UPLOAD_REQUIRED"
        else:
            status = "NOT_STARTED"
        updates.append((status, count, max(dates) if dates else None, now, slot["package_slot_instance_id"]))
        slot.update({"completion_status": status, "selected_document_count": count, "latest_selected_document_date": max(dates) if dates else None})
    with database.get_connection(db_path) as connection:
        connection.executemany(
            "UPDATE package_slot_instances SET completion_status=?, selected_document_count=?, latest_selected_document_date=?, updated_at=? WHERE package_slot_instance_id=?",
            updates,
        )
    summary = completion_summary(slots)
    instance = get_package_recipe_instance(package_id, db_path=db_path)
    if instance:
        target = "READY_FOR_REVIEW" if summary["readiness"] in {"READY_FOR_REVIEW", "READY_WITH_ACKNOWLEDGEMENTS"} else "IN_PROGRESS"
        if instance["status"] != target:
            with database.get_connection(db_path) as connection:
                connection.execute("UPDATE package_recipe_instances SET status=? WHERE package_id=?", (target, package_id))
            _audit("PACKAGE_READINESS_CHANGED", actor=actor, package_id=package_id, recipe_id=instance["recipe_id"],
                   details={"from": instance["status"], "to": target, "readiness": summary["readiness"]}, db_path=db_path)
    return summary


def completion_summary(slots: list[dict[str, Any]]) -> dict[str, Any]:
    applicable = [slot for slot in slots if slot["completion_status"] != "NOT_APPLICABLE"]
    complete = [slot for slot in applicable if slot["completion_status"] == "COMPLETE"]
    required = [slot for slot in applicable if slot["requirement_snapshot"] == "REQUIRED"]
    recommended = [slot for slot in applicable if slot["requirement_snapshot"] == "RECOMMENDED"]
    optional = [slot for slot in applicable if slot["requirement_snapshot"] in {"OPTIONAL", "MANUAL_ONLY", "CONDITIONAL"}]
    required_open = [slot for slot in required if slot["completion_status"] != "COMPLETE"]
    required_unavailable = [slot for slot in required_open if slot["completion_status"] == "NOT_AVAILABLE" and slot.get("analyst_notes")]
    blockers = [slot for slot in required_open if slot not in required_unavailable]
    if not required_open:
        readiness = "READY_FOR_REVIEW"
    elif not blockers and len(required_unavailable) == len(required_open):
        readiness = "READY_WITH_ACKNOWLEDGEMENTS"
    else:
        readiness = "NOT_READY"
    return {
        "overall_complete": len(complete),
        "overall_total": len(applicable),
        "overall_percent": round(100 * len(complete) / max(1, len(applicable))),
        "required_complete": sum(slot["completion_status"] == "COMPLETE" for slot in required),
        "required_total": len(required),
        "recommended_complete": sum(slot["completion_status"] == "COMPLETE" for slot in recommended),
        "recommended_total": len(recommended),
        "optional_complete": sum(slot["completion_status"] == "COMPLETE" for slot in optional),
        "optional_total": len(optional),
        "manual_uploads_required": sum(slot["completion_status"] in {"MANUAL_UPLOAD_REQUIRED", "MISSING"} for slot in applicable),
        "needs_review": sum(slot["completion_status"] in {"NEEDS_ANALYST_REVIEW", "PARTIAL"} for slot in applicable),
        "acknowledged_unavailable": sum(slot["completion_status"] == "NOT_AVAILABLE" for slot in applicable),
        "readiness": readiness,
    }


def package_guidance(slots: list[dict[str, Any]]) -> str:
    priorities = (
        (lambda slot: slot["requirement_snapshot"] == "REQUIRED" and slot["completion_status"] in {"NEEDS_ANALYST_REVIEW", "PARTIAL"}, "Review the proposed assignment for {name}."),
        (lambda slot: slot["requirement_snapshot"] == "REQUIRED" and slot["completion_status"] in {"MANUAL_UPLOAD_REQUIRED", "MISSING"}, "Upload {name}, or acknowledge that it is unavailable."),
        (lambda slot: slot["requirement_snapshot"] == "RECOMMENDED" and slot["completion_status"] == "MANUAL_UPLOAD_REQUIRED", "Upload the recommended item: {name}."),
        (lambda slot: slot["requirement_snapshot"] in {"OPTIONAL", "CONDITIONAL", "MANUAL_ONLY"} and slot["completion_status"] == "NEEDS_ANALYST_REVIEW", "Review applicability for {name}."),
    )
    for predicate, message in priorities:
        match = next((slot for slot in slots if predicate(slot)), None)
        if match:
            return message.format(name=match["display_name_snapshot"])
    return "All required slots are complete. Review the package before locking."


CLASSIFICATION_RULES: tuple[tuple[str, tuple[str, ...], float, bool], ...] = (
    ("BBG-FA Credit Ratios", ("credit ratios", "credit-ratios"), 0.99, False),
    ("DRSK Default Risk", ("drsk",), 0.99, False),
    ("BBG-ANR", (" anr ", "_anr_", "-anr-"), 0.98, False),
    ("BBG-FA", (" fa ", "_fa_", "-fa-"), 0.92, False),
    ("Initiated Coverage Report", ("initiation", "initiated coverage"), 0.96, False),
    ("Industry Report", ("industry",), 0.95, False),
    ("Credit Reports", ("moody", "gimme credit", "credit report", "s&p credit"), 0.95, False),
    ("Latest Earnings Call Transcript", ("transcript", "earnings commentary", "earnings-commentary"), 0.78, True),
    ("Available Supplemental or Earnings Presentation", ("earnings presentation", "supplemental"), 0.93, False),
    ("Morningstar Report and Most Recent Model", ("morningstar",), 0.95, False),
    ("Sell-Side Reports", ("jpm", "goldman", " gs ", "wfc", "ubs", "evercore", "jefferies", "deutsche", "needham", "mizuho"), 0.88, False),
)


def classify_filename(filename: str, slots: list[dict[str, Any]]) -> dict[str, Any]:
    normalized = f" {re.sub(r'[^a-z0-9]+', ' ', Path(filename).stem.lower())} "
    slot_by_name = {slot.get("display_name_snapshot") or slot.get("display_name"): slot for slot in slots}
    for name, tokens, confidence, ambiguous in CLASSIFICATION_RULES:
        matched = [token.strip() for token in tokens if token in normalized]
        if matched and name in slot_by_name:
            slot = slot_by_name[name]
            return {
                "suggested_slot_id": slot.get("package_slot_instance_id") or slot.get("slot_id"),
                "suggested_slot_name": name,
                "confidence": confidence,
                "requires_review": ambiguous or confidence < 0.9,
                "matched_tokens": matched,
                "reason": f"Filename matched deterministic token(s): {', '.join(matched)}.",
            }
    return {"suggested_slot_id": None, "suggested_slot_name": None, "confidence": 0.0, "requires_review": True,
            "matched_tokens": [], "reason": "No deterministic Phase 6A filename rule matched."}


def suggest_document_assignments(package_id: str, document_ids: list[str], *, actor: str, db_path: Path | str = config.DATABASE_PATH) -> list[dict[str, Any]]:
    slots = list_slot_instances(package_id, db_path=db_path)
    documents = {row["document_id"]: row for row in database.list_documents_by_package(package_id, db_path=db_path)}
    suggestions: list[dict[str, Any]] = []
    for document_id in document_ids:
        document = documents.get(document_id)
        if not document:
            continue
        filename = document.get("original_filename") or document.get("local_filename") or document.get("title") or ""
        suggestion = classify_filename(filename, slots)
        suggestion.update({"document_id": document_id, "filename": filename})
        suggestions.append(suggestion)
        if suggestion["suggested_slot_id"]:
            try:
                assignment = assign_document(
                    suggestion["suggested_slot_id"], document_id, actor=actor, assignment_source="FILENAME_RULE",
                    status="NEEDS_REVIEW" if suggestion["requires_review"] else "SUGGESTED", selected=False,
                    suggestion=suggestion, db_path=db_path,
                )
                suggestion["assignment_id"] = assignment["assignment_id"]
            except ValueError:
                pass
    return suggestions


def assign_document(
    slot_instance_id: str,
    document_id: str,
    *,
    actor: str,
    assignment_source: str = "ANALYST",
    status: str = "APPROVED",
    selected: bool = True,
    suggestion: dict[str, Any] | None = None,
    override_cap: bool = False,
    override_reason: str = "",
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    database.initialize_database(db_path)
    now = database.utc_now_iso()
    suggestion = suggestion or {}
    with database.get_connection(db_path) as connection:
        slot = connection.execute("SELECT * FROM package_slot_instances WHERE package_slot_instance_id=?", (slot_instance_id,)).fetchone()
        document = connection.execute("SELECT * FROM documents WHERE document_id=?", (document_id,)).fetchone()
        if not slot or not document or slot["package_id"] != document["package_id"]:
            raise ValueError("The slot and document must belong to the same package.")
        existing = connection.execute(
            "SELECT * FROM slot_document_assignments WHERE package_slot_instance_id=? AND document_id=? AND assignment_status NOT IN ('REJECTED','REPLACED','REMOVED')",
            (slot_instance_id, document_id),
        ).fetchone()
        if existing:
            return dict(existing)
        another_slot = connection.execute(
            "SELECT 1 FROM slot_document_assignments WHERE document_id=? AND package_slot_instance_id!=? AND assignment_status NOT IN ('REJECTED','REPLACED','REMOVED') LIMIT 1",
            (document_id, slot_instance_id),
        ).fetchone()
        if another_slot and assignment_source != "GENERATED_DERIVATIVE":
            raise ValueError("This original document is already assigned to another slot. Create a recorded generated derivative for multi-slot use.")
        approved_count = int(connection.execute(
            "SELECT COUNT(*) FROM slot_document_assignments WHERE package_slot_instance_id=? AND selected_for_package=1 AND assignment_status='APPROVED'",
            (slot_instance_id,),
        ).fetchone()[0])
        if selected and status == "APPROVED" and approved_count >= int(slot["maximum_documents"]) and not override_cap:
            raise ValueError("This slot has reached its document cap. Replace a document or record an approved override.")
        if override_cap and not override_reason.strip():
            raise ValueError("A reason is required for a slot-cap override.")
        assignment_id = _token("ASG")
        connection.execute(
            """INSERT INTO slot_document_assignments(
                assignment_id, package_slot_instance_id, package_id, document_id, assignment_source,
                suggested_slot_id, final_slot_id, suggestion_confidence, suggestion_reason, matched_tokens_json,
                assignment_status, selected_for_package, highlighted_research, display_order, analyst_notes,
                assigned_at, assigned_by, approved_at, approved_by, replaced_assignment_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, NULL)""",
            (assignment_id, slot_instance_id, slot["package_id"], document_id, assignment_source,
             suggestion.get("suggested_slot_id"), slot["slot_id"], suggestion.get("confidence"), suggestion.get("reason"),
             _json(suggestion.get("matched_tokens", [])), status, int(selected), approved_count,
             override_reason or None, now, actor, now if status == "APPROVED" else None, actor if status == "APPROVED" else None),
        )
        if override_cap:
            connection.execute("UPDATE package_slot_instances SET cap_override_approved=1 WHERE package_slot_instance_id=?", (slot_instance_id,))
        result = connection.execute("SELECT * FROM slot_document_assignments WHERE assignment_id=?", (assignment_id,)).fetchone()
    _audit("ASSIGNMENT_APPROVED" if status == "APPROVED" else "ASSIGNMENT_SUGGESTED", actor=actor,
           package_id=slot["package_id"], slot_instance_id=slot_instance_id, document_id=document_id,
           details={"assignment_id": assignment_id, "confidence": suggestion.get("confidence"), "override_cap": override_cap, "override_reason": override_reason}, db_path=db_path)
    recalculate_completion(slot["package_id"], actor=actor, db_path=db_path)
    return dict(result)


def update_assignment(assignment_id: str, action: str, *, actor: str, notes: str = "", db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    mapping = {"approve": ("APPROVED", 1), "reject": ("REJECTED", 0), "remove": ("REMOVED", 0)}
    if action not in mapping:
        raise ValueError("Unsupported assignment action.")
    status, selected = mapping[action]
    now = database.utc_now_iso()
    with database.get_connection(db_path) as connection:
        row = connection.execute("SELECT * FROM slot_document_assignments WHERE assignment_id=?", (assignment_id,)).fetchone()
        if not row:
            raise ValueError("Assignment does not exist.")
        if action == "approve":
            slot = connection.execute("SELECT * FROM package_slot_instances WHERE package_slot_instance_id=?", (row["package_slot_instance_id"],)).fetchone()
            count = int(connection.execute("SELECT COUNT(*) FROM slot_document_assignments WHERE package_slot_instance_id=? AND selected_for_package=1 AND assignment_status='APPROVED'", (row["package_slot_instance_id"],)).fetchone()[0])
            if count >= int(slot["maximum_documents"]):
                raise ValueError("This slot has reached its document cap.")
        connection.execute(
            "UPDATE slot_document_assignments SET assignment_status=?, selected_for_package=?, analyst_notes=?, approved_at=?, approved_by=? WHERE assignment_id=?",
            (status, selected, notes or row["analyst_notes"], now if status == "APPROVED" else None, actor if status == "APPROVED" else None, assignment_id),
        )
        refreshed = connection.execute("SELECT * FROM slot_document_assignments WHERE assignment_id=?", (assignment_id,)).fetchone()
    _audit("ASSIGNMENT_APPROVED" if action == "approve" else "ASSIGNMENT_REJECTED", actor=actor,
           package_id=row["package_id"], slot_instance_id=row["package_slot_instance_id"], document_id=row["document_id"],
           details={"assignment_id": assignment_id, "action": action}, db_path=db_path)
    recalculate_completion(row["package_id"], actor=actor, db_path=db_path)
    return dict(refreshed)


def replace_assignment(
    assignment_id: str,
    replacement_document_id: str,
    *,
    actor: str,
    reason: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    if not reason.strip():
        raise ValueError("A replacement reason is required.")
    now = database.utc_now_iso()
    replacement_id = _token("ASG")
    with database.get_connection(db_path) as connection:
        old = connection.execute("SELECT * FROM slot_document_assignments WHERE assignment_id=?", (assignment_id,)).fetchone()
        document = connection.execute("SELECT * FROM documents WHERE document_id=?", (replacement_document_id,)).fetchone()
        if not old or not document or old["package_id"] != document["package_id"]:
            raise ValueError("Replacement assignment and document must belong to the same package.")
        duplicate = connection.execute(
            "SELECT 1 FROM slot_document_assignments WHERE package_slot_instance_id=? AND document_id=? AND assignment_status NOT IN ('REJECTED','REPLACED','REMOVED')",
            (old["package_slot_instance_id"], replacement_document_id),
        ).fetchone()
        if duplicate:
            raise ValueError("The replacement document is already assigned to this slot.")
        another_slot = connection.execute(
            "SELECT 1 FROM slot_document_assignments WHERE document_id=? AND package_slot_instance_id!=? AND assignment_status NOT IN ('REJECTED','REPLACED','REMOVED') LIMIT 1",
            (replacement_document_id, old["package_slot_instance_id"]),
        ).fetchone()
        if another_slot:
            raise ValueError("The replacement original is already assigned to another slot.")
        connection.execute(
            "UPDATE slot_document_assignments SET assignment_status='REPLACED', selected_for_package=0 WHERE assignment_id=?",
            (assignment_id,),
        )
        connection.execute(
            """INSERT INTO slot_document_assignments(
                assignment_id, package_slot_instance_id, package_id, document_id, assignment_source,
                suggested_slot_id, final_slot_id, suggestion_confidence, suggestion_reason, matched_tokens_json,
                assignment_status, selected_for_package, highlighted_research, display_order, analyst_notes,
                assigned_at, assigned_by, approved_at, approved_by, replaced_assignment_id
            ) VALUES (?, ?, ?, ?, 'ANALYST_REPLACEMENT', NULL, ?, NULL, ?, '[]', 'APPROVED', 1, 0, ?, ?, ?, ?, ?, ?, ?)""",
            (replacement_id, old["package_slot_instance_id"], old["package_id"], replacement_document_id,
             old["final_slot_id"], reason.strip(), old["display_order"], reason.strip(), now, actor, now, actor, assignment_id),
        )
        replacement = connection.execute("SELECT * FROM slot_document_assignments WHERE assignment_id=?", (replacement_id,)).fetchone()
    _audit("ASSIGNMENT_REPLACED", actor=actor, package_id=old["package_id"], slot_instance_id=old["package_slot_instance_id"],
           document_id=replacement_document_id, details={"replaced_assignment_id": assignment_id, "reason": reason}, db_path=db_path)
    recalculate_completion(old["package_id"], actor=actor, db_path=db_path)
    return dict(replacement)


def mark_slot(
    slot_instance_id: str,
    status: str,
    *,
    reason: str,
    actor: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    if status not in {"NOT_AVAILABLE", "NOT_APPLICABLE", "RESTORE"}:
        raise ValueError("Unsupported slot status.")
    if status != "RESTORE" and not reason.strip():
        raise ValueError("A reason is required.")
    with database.get_connection(db_path) as connection:
        slot = connection.execute("SELECT * FROM package_slot_instances WHERE package_slot_instance_id=?", (slot_instance_id,)).fetchone()
        if not slot:
            raise ValueError("Slot does not exist.")
        applicability = "NOT_APPLICABLE" if status == "NOT_APPLICABLE" else "APPLICABLE"
        completion = "NOT_APPLICABLE" if status == "NOT_APPLICABLE" else "NOT_AVAILABLE" if status == "NOT_AVAILABLE" else "NOT_STARTED"
        connection.execute(
            "UPDATE package_slot_instances SET applicability_status=?, completion_status=?, analyst_acknowledged=?, analyst_notes=?, updated_at=? WHERE package_slot_instance_id=?",
            (applicability, completion, int(status != "RESTORE"), reason.strip() or None, database.utc_now_iso(), slot_instance_id),
        )
        refreshed = connection.execute("SELECT * FROM package_slot_instances WHERE package_slot_instance_id=?", (slot_instance_id,)).fetchone()
    event = "SLOT_RESTORED" if status == "RESTORE" else "SLOT_MARKED_NOT_APPLICABLE" if status == "NOT_APPLICABLE" else "SLOT_MARKED_UNAVAILABLE"
    _audit(event, actor=actor, package_id=slot["package_id"], slot_instance_id=slot_instance_id, details={"reason": reason}, db_path=db_path)
    recalculate_completion(slot["package_id"], actor=actor, db_path=db_path)
    return dict(refreshed)


def add_slot_note(slot_instance_id: str, note: str, *, actor: str, db_path: Path | str = config.DATABASE_PATH) -> None:
    with database.get_connection(db_path) as connection:
        slot = connection.execute("SELECT * FROM package_slot_instances WHERE package_slot_instance_id=?", (slot_instance_id,)).fetchone()
        if not slot:
            raise ValueError("Slot does not exist.")
        original = slot["analyst_notes"]
        connection.execute("UPDATE package_slot_instances SET analyst_notes=?, updated_at=? WHERE package_slot_instance_id=?", (note.strip() or None, database.utc_now_iso(), slot_instance_id))
        connection.execute(
            "INSERT INTO recipe_corrections VALUES (?, ?, ?, NULL, 'SLOT_NOTE', ?, ?, ?, ?, ?)",
            (_token("COR"), slot["package_id"], slot_instance_id, original, note.strip(), "Analyst note updated.", database.utc_now_iso(), actor),
        )


def set_highlighted(assignment_id: str, highlighted: bool, *, actor: str, db_path: Path | str = config.DATABASE_PATH) -> None:
    with database.get_connection(db_path) as connection:
        row = connection.execute("SELECT * FROM slot_document_assignments WHERE assignment_id=?", (assignment_id,)).fetchone()
        if not row or row["assignment_status"] != "APPROVED":
            raise ValueError("Only approved assignments can be highlighted.")
        connection.execute("UPDATE slot_document_assignments SET highlighted_research=? WHERE assignment_id=?", (int(highlighted), assignment_id))
    _audit("DOCUMENT_HIGHLIGHTED", actor=actor, package_id=row["package_id"], slot_instance_id=row["package_slot_instance_id"],
           document_id=row["document_id"], details={"highlighted": highlighted}, db_path=db_path)


def board_payload(package_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    started = datetime.now().timestamp()
    package = database.get_package_by_package_id(package_id, db_path=db_path)
    if not package:
        raise ValueError("Package does not exist.")
    instance = get_package_recipe_instance(package_id, db_path=db_path)
    if not instance:
        return {"package": package, "legacy": True, "load_ms": round((datetime.now().timestamp() - started) * 1000, 1)}
    summary = recalculate_completion(package_id, db_path=db_path)
    slots = list_slot_instances(package_id, db_path=db_path)
    assignments = list_assignments(package_id, db_path=db_path)
    by_slot: dict[str, list[dict[str, Any]]] = {}
    for assignment in assignments:
        by_slot.setdefault(assignment["package_slot_instance_id"], []).append(assignment)
    recipe_meta = instance["recipe_snapshot"].get("recipe", {})
    return {
        "package": package, "legacy": False, "instance": instance, "recipe": recipe_meta,
        "slots": slots, "assignments": assignments, "assignments_by_slot": by_slot,
        "summary": summary, "guidance": package_guidance(slots),
        "highlighted": [item for item in assignments if item["assignment_status"] == "APPROVED" and item["highlighted_research"]],
        "load_ms": round((datetime.now().timestamp() - started) * 1000, 1),
    }


def update_package_header(package_id: str, *, compilation_date: date, research_cutoff: date, compiled_by: str, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    with database.get_connection(db_path) as connection:
        connection.execute(
            "UPDATE packages SET compilation_date=?, research_cutoff_date=?, compiled_by=?, updated_at=? WHERE package_id=?",
            (compilation_date.isoformat(), research_cutoff.isoformat(), compiled_by.strip(), database.utc_now_iso(), package_id),
        )
    return database.get_package_by_package_id(package_id, db_path=db_path) or {}


def export_checklist_xlsx(package_id: str, *, actor: str, db_path: Path | str = config.DATABASE_PATH) -> bytes:
    payload = board_payload(package_id, db_path=db_path)
    if payload.get("legacy"):
        raise ValueError("Checklist export requires a Phase 6 recipe instance.")
    package, recipe = payload["package"], payload["recipe"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Package Checklist"
    sheet.freeze_panes = "A9"
    sheet.merge_cells("A1:I1")
    sheet["A1"] = "Comprehensive Equity Research Package"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1B1F23")
    sheet["A1"].alignment = Alignment(horizontal="center")
    metadata = (
        ("Company", package.get("company_name")), ("Ticker", package.get("ticker")),
        ("Compilation date", package.get("compilation_date")), ("Research cutoff", package.get("research_cutoff_date")),
        ("Next earnings date", "Not available"), ("Next ex-dividend date", "Not available"),
        ("Compiled by", package.get("compiled_by")), ("Recipe", f"{recipe.get('recipe_name')} v{recipe.get('version')}"),
        ("Package ID", package_id),
    )
    for index, (label, value) in enumerate(metadata, start=2):
        column = 1 if index < 7 else 5
        row = index if index < 7 else index - 5
        sheet.cell(row, column, label).font = Font(bold=True)
        sheet.cell(row, column + 1, value or "")
    headers = ["Order", "Research Item", "Section", "Requirement", "Status", "Selected Document", "Source", "Document Date", "Notes"]
    header_row = 8
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7A1F27")
    for row_index, slot in enumerate(payload["slots"], start=9):
        assignments = [item for item in payload["assignments_by_slot"].get(slot["package_slot_instance_id"], []) if item["assignment_status"] == "APPROVED"]
        order = "" if slot["order_number"] is None else str(slot["order_number"])
        if slot["suborder"]:
            order = f"{order}.{slot['suborder']} supplemental"
        selected = "; ".join(item.get("document_title") or item.get("original_filename") or item["document_id"] for item in assignments)
        sources = "; ".join(item.get("source_name") or "" for item in assignments)
        dates = "; ".join(str(item.get("document_date") or item.get("publication_date") or "") for item in assignments)
        values = [order, slot["display_name_snapshot"], slot["section_snapshot"], slot["requirement_snapshot"],
                  slot["completion_status"], selected, sources, dates, slot.get("analyst_notes") or ""]
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, value)
    start = 10 + len(payload["slots"])
    sheet.cell(start, 1, "Highlighted Research").font = Font(bold=True, color="FFFFFF")
    sheet.cell(start, 1).fill = PatternFill("solid", fgColor="1B1F23")
    for offset, item in enumerate(payload["highlighted"], 1):
        sheet.cell(start + offset, 1, item.get("document_title") or item["document_id"])
        sheet.cell(start + offset, 2, item.get("source_name") or "")
    widths = (14, 42, 34, 18, 24, 42, 24, 16, 42)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = io.BytesIO()
    workbook.save(output)
    _audit("CHECKLIST_EXPORTED", actor=actor, package_id=package_id, recipe_id=payload["instance"]["recipe_id"],
           details={"slot_count": len(payload["slots"]), "sha256": hashlib.sha256(output.getvalue()).hexdigest()}, db_path=db_path)
    return output.getvalue()


def export_package_snapshot(package_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> bytes:
    payload = board_payload(package_id, db_path=db_path)
    if payload.get("legacy"):
        raise ValueError("Snapshot export requires a Phase 6 recipe instance.")
    documents = database.list_documents_by_package(package_id, db_path=db_path)
    safe_documents = [
        {key: value for key, value in document.items() if key not in {"local_path", "error_message", "authorization_confirmed"}}
        for document in documents
    ]
    with database.get_connection(db_path) as connection:
        corrections = [dict(row) for row in connection.execute("SELECT * FROM recipe_corrections WHERE package_id=? ORDER BY corrected_at", (package_id,)).fetchall()]
    snapshot = {
        "snapshot_version": "6A.1", "package_metadata": payload["package"], "recipe_metadata": payload["recipe"],
        "recipe_snapshot": payload["instance"]["recipe_snapshot"], "slot_instances": payload["slots"],
        "assignment_metadata": payload["assignments"], "document_metadata": safe_documents,
        "correction_history": corrections, "completion_summary": payload["summary"],
    }
    return _json(snapshot).encode("utf-8")


def clone_legacy_package(
    source_package_id: str,
    *,
    created_by: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    """Clone legacy metadata/doc records into a distinct recipe draft without changing the source."""
    source = database.get_package_by_package_id(source_package_id, db_path=db_path)
    if not source:
        raise ValueError("Legacy package does not exist.")
    if get_package_recipe_instance(source_package_id, db_path=db_path):
        raise ValueError("Only a legacy package can be cloned into a Phase 6 draft.")
    try:
        cutoff = date.fromisoformat(str(source["research_cutoff_date"])[:10])
    except ValueError:
        cutoff = date.today()
    cloned = create_package_from_active_recipe(
        source,
        research_cutoff=cutoff,
        compilation_date=date.today(),
        compiled_by=created_by,
        created_by=created_by,
        db_path=db_path,
    )
    now = database.utc_now_iso()
    copied_ids: list[str] = []
    with database.get_connection(db_path) as connection:
        connection.execute(
            "UPDATE packages SET source_legacy_package_id=? WHERE package_id=?",
            (source_package_id, cloned["package_id"]),
        )
        columns = [row["name"] for row in connection.execute("PRAGMA table_info(documents)").fetchall() if row["name"] != "id"]
        source_docs = connection.execute("SELECT * FROM documents WHERE package_id=?", (source_package_id,)).fetchall()
        for document in source_docs:
            copied = {column: document[column] for column in columns}
            copied["document_id"] = database.generate_document_id("DOC-CLONE")
            copied["package_id"] = cloned["package_id"]
            copied["ticker"] = cloned["ticker"]
            copied["created_at"] = now
            copied["updated_at"] = now
            copied["archive_origin_document_id"] = document["document_id"]
            connection.execute(
                f"INSERT INTO documents({', '.join(columns)}) VALUES ({', '.join('?' for _ in columns)})",
                tuple(copied[column] for column in columns),
            )
            copied_ids.append(copied["document_id"])
    suggest_document_assignments(cloned["package_id"], copied_ids, actor=created_by, db_path=db_path)
    _audit("LEGACY_PACKAGE_CLONED", actor=created_by, package_id=cloned["package_id"],
           details={"source_legacy_package_id": source_package_id, "copied_document_records": len(copied_ids)}, db_path=db_path)
    return database.get_package_by_package_id(cloned["package_id"], db_path=db_path) or cloned


def import_package_snapshot(
    content: bytes,
    *,
    imported_by: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    """Create a distinct draft from safe snapshot metadata; document bytes remain deferred."""
    if len(content) > 10 * 1024 * 1024:
        raise ValueError("Snapshot exceeds the Phase 6A metadata import limit.")
    try:
        snapshot = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("Snapshot is not valid UTF-8 JSON.") from exc
    forbidden = {"api_key", "authorization", "openai_prompt", "raw_document_text", "secret"}

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            if any(str(key).lower() in forbidden for key in value):
                raise ValueError("Snapshot contains a forbidden secret or raw-text field.")
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(snapshot)
    metadata = snapshot.get("package_metadata") or {}
    try:
        cutoff = date.fromisoformat(str(metadata.get("research_cutoff_date") or date.today().isoformat())[:10])
    except ValueError:
        cutoff = date.today()
    new_package = create_package_from_active_recipe(
        metadata,
        research_cutoff=cutoff,
        compilation_date=date.today(),
        compiled_by=imported_by,
        created_by=imported_by,
        db_path=db_path,
    )
    digest = hashlib.sha256(content).hexdigest()
    with database.get_connection(db_path) as connection:
        connection.execute(
            "INSERT INTO package_snapshot_imports VALUES (?, ?, ?, ?, ?, ?)",
            (_token("SNAP"), str(metadata.get("package_id") or "UNKNOWN"), new_package["package_id"], digest,
             database.utc_now_iso(), imported_by),
        )
    return new_package


def database_audit_details(*, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    path = Path(db_path).resolve()
    configured = Path(config.DATABASE_PATH).resolve()
    if path != configured:
        environment, persistent = "TEST", False
    else:
        environment = config.DATABASE_ENVIRONMENT
        persistent = environment in {"DEVELOPMENT", "STREAMLIT_CLOUD"}
    database.initialize_database(db_path)
    with database.get_connection(db_path) as connection:
        count = int(connection.execute("SELECT COUNT(*) FROM packages").fetchone()[0])
        recent = connection.execute("SELECT package_id FROM packages ORDER BY created_at DESC LIMIT 1").fetchone()
        version = connection.execute("SELECT schema_value FROM schema_metadata WHERE schema_key='database_schema_version'").fetchone()
    return {
        "environment": environment, "storage": "Persistent" if persistent else "Temporary",
        "package_count": count, "most_recent_package_id": recent[0] if recent else "None",
        "schema_version": version[0] if version else "Unknown",
    }
