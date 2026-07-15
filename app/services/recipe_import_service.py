from __future__ import annotations

import hashlib
import json
import re
import secrets
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app import config
from app.utils import database


IMPORTER_VERSION = "6A.1"
COMMON_EQUITY_SHEETS = ("Template", "Instructions", "MDT")
RECIPE_NAME = "Cutler Common Equity Research Package"

SECTION_ORDER = (
    ("COMPANY_SNAPSHOT", "Company Snapshot"),
    ("LICENSED_RESEARCH", "Licensed and Third-Party Research"),
    ("COMPANY_MATERIALS", "Earnings and Company Materials"),
    ("SEC_FILINGS", "SEC Filings and Extracted Sections"),
    ("INTERNAL_ANALYSIS", "Internal Cutler Analysis"),
    ("FINAL_MEMO", "Final Memo"),
)

CANONICAL_NAMES: tuple[tuple[tuple[str, ...], str], ...] = (
    (("bbg des",), "BBG-DES"),
    (("bbg dvd",), "BBG-DVD"),
    (("bbg ddis",), "BBG-DDIS"),
    (("bbg hds",), "BBG-HDS"),
    (("hoover",), "Hoover Report"),
    (("morningstar",), "Morningstar Report and Most Recent Model"),
    (("bbg anr",), "BBG-ANR"),
    (("sell side downgrade",), "Sell-Side Downgrade"),
    (("sell side",), "Sell-Side Reports"),
    (("credit report",), "Credit Reports"),
    (("initiated coverage", "initiation"), "Initiated Coverage Report"),
    (("independent short",), "Independent Short-Sale Report"),
    (("industry report",), "Industry Report"),
    (("foreign country",), "Foreign Country Economic Report"),
    (("foreign currency",), "Foreign Currency Forecast"),
    (("press release",), "Material Company Press Releases Since Last Earnings Release"),
    (("earnings release",), "Latest Earnings Release"),
    (("supplemental",), "Available Supplemental or Earnings Presentation"),
    (("call audio",), "Latest Earnings Call Audio"),
    (("call transcript", "earnings commentary"), "Latest Earnings Call Transcript"),
    (("investor presentation", "company presentation"), "Investor Presentations"),
    (("liquidity and capital",), "Liquidity and Capital Resources"),
    (("description of business",), "Description of Business and Risk"),
    (("properties", "locations"), "Properties or Locations for REITs"),
    (("executive compensation",), "Executive Compensation Information"),
    (("financial statement",), "Financial Statements from the Latest Filing"),
    (("most recent 10 q",), "Most Recent 10-Q and 10-K"),
    (("cast",), "CAST Summary Chart"),
    (("drsk",), "DRSK Default Risk"),
    (("ccm historical",), "CCM Historical Multiples Valuation"),
    (("credit ratio",), "BBG-FA Credit Ratios"),
    (("bbg fa",), "BBG-FA"),
)

REQUIRED_NAMES = {
    "Latest Earnings Release",
    "Latest Earnings Call Transcript",
    "Liquidity and Capital Resources",
    "Description of Business and Risk",
    "Most Recent 10-Q and 10-K",
}


def _token() -> str:
    return secrets.token_hex(8).upper()


def _json(value: Any) -> str:
    return json.dumps(value, sort_keys=True, default=str)


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _clean(value).lower()).strip()


def _canonical_name(raw_name: str) -> str:
    key = _key(raw_name)
    for terms, canonical in CANONICAL_NAMES:
        if any(term in key for term in terms):
            return canonical
    return _clean(raw_name)


def _section_for(name: str) -> tuple[str, str]:
    key = _key(name)
    if any(token in key for token in ("bbg des", "bbg dvd", "bbg ddis", "bbg hds", "hoover", "morningstar", "bbg anr")):
        return SECTION_ORDER[0]
    if any(token in key for token in ("sell side", "credit report", "initiated", "short sale", "industry", "foreign country", "foreign currency")):
        return SECTION_ORDER[1]
    if any(token in key for token in ("earnings release", "supplemental", "call audio", "transcript", "presentation", "press release")):
        return SECTION_ORDER[2]
    if any(token in key for token in ("liquidity", "business and risk", "properties", "executive compensation", "financial statement", "10 q and 10 k")):
        return SECTION_ORDER[3]
    if any(token in key for token in ("cast", "drsk", "ccm", "bbg fa")):
        return SECTION_ORDER[4]
    return SECTION_ORDER[5]


def _requirement_for(name: str, raw_requirement: str, order_number: int | None) -> str:
    if name in REQUIRED_NAMES:
        return "REQUIRED"
    if name in {"Sell-Side Downgrade", "Independent Short-Sale Report", "Properties or Locations for REITs"}:
        return "CONDITIONAL"
    if order_number is None or name in {"Latest Earnings Call Audio", "CAST Summary Chart"}:
        return "MANUAL_ONLY"
    if _key(raw_requirement) in {"n a", "na"}:
        return "CONDITIONAL"
    return "RECOMMENDED"


def _merged_value(sheet: Any, row: int, column: int) -> Any:
    value = sheet.cell(row, column).value
    if value is not None:
        return value
    coordinate = sheet.cell(row, column).coordinate
    for merged in sheet.merged_cells.ranges:
        if coordinate in merged:
            return sheet.cell(merged.min_row, merged.min_col).value
    return None


def _raw_rows(sheet: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_number in range(1, sheet.max_row + 1):
        values = [_merged_value(sheet, row_number, column) for column in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        coordinates = [f"{get_column_letter(column)}{row_number}" for column in range(1, sheet.max_column + 1)]
        rows.append({"row_number": row_number, "coordinates": coordinates, "values": values})
    return rows


def _source_catalog(workbook: Any) -> dict[str, dict[str, dict[str, Any]]]:
    layouts = {
        "Template": (5, 6, None, 7),
        "Instructions": (2, 3, 4, 6),
        "MDT": (4, 6, 7, None),
    }
    catalog: dict[str, dict[str, dict[str, Any]]] = {}
    for sheet_name, (name_col, requirement_col, source_col, instruction_col) in layouts.items():
        sheet = workbook[sheet_name]
        sheet_items: dict[str, dict[str, Any]] = {}
        for row in range(1, sheet.max_row + 1):
            name = _clean(_merged_value(sheet, row, name_col))
            canonical = _canonical_name(name)
            if not name or canonical == name and canonical not in {item[1] for item in CANONICAL_NAMES}:
                continue
            sheet_items.setdefault(
                canonical,
                {
                    "row": row,
                    "name": name,
                    "requirement": _clean(_merged_value(sheet, row, requirement_col)),
                    "source": _clean(_merged_value(sheet, row, source_col)) if source_col else "",
                    "instructions": _clean(_merged_value(sheet, row, instruction_col)) if instruction_col else "",
                },
            )
        catalog[sheet_name] = sheet_items
    return catalog


def inspect_common_equity_workbook(workbook_path: Path | str | None = None) -> dict[str, Any]:
    """Inspect the workbook without persisting data or changing the source file."""
    path = Path(workbook_path or config.RECIPE_WORKBOOK_PATH).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError("The configured recipe workbook was not found.")
    source_bytes = path.read_bytes()
    workbook = load_workbook(path, data_only=False, read_only=False)
    available = list(workbook.sheetnames)
    missing = [name for name in COMMON_EQUITY_SHEETS if name not in available]
    if missing:
        raise ValueError(f"Workbook is missing required source sheets: {', '.join(missing)}")
    catalog = _source_catalog(workbook)
    raw_rows = {name: _raw_rows(workbook[name]) for name in COMMON_EQUITY_SHEETS}

    mdt = workbook["MDT"]
    slots: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    previous_order: int | None = None
    suborders: dict[int, int] = {}
    seen_names: dict[str, int] = {}
    for row in range(1, mdt.max_row + 1):
        raw_name = _clean(_merged_value(mdt, row, 4))
        canonical = _canonical_name(raw_name)
        recognized = canonical in {item[1] for item in CANONICAL_NAMES}
        if not raw_name or not recognized:
            continue
        raw_order = _merged_value(mdt, row, 3)
        try:
            order_number = int(raw_order) if raw_order not in (None, "") else None
        except (TypeError, ValueError):
            order_number = None
        if order_number is None:
            order_number = previous_order
            suborder = suborders.get(order_number or 0, 0) + 1
            suborders[order_number or 0] = suborder
            warnings.append({"sheet": "MDT", "row": row, "code": "UNNUMBERED_SUPPLEMENTAL", "message": raw_name})
        else:
            previous_order = order_number
            suborder = 0
        raw_requirement = _clean(_merged_value(mdt, row, 6))
        preferred_source = _clean(_merged_value(mdt, row, 7))
        section_code, section_name = _section_for(canonical)
        source_coordinates = {
            "order": f"C{row}", "name": f"D{row}", "requirement": f"F{row}", "preferred_source": f"G{row}"
        }
        warning_codes = ["AMBIGUOUS_REQUIREMENT"] if raw_requirement.upper() in {"Y", "N", "N/A", "-", ""} else []
        if warning_codes:
            warnings.append({"sheet": "MDT", "row": row, "code": warning_codes[0], "message": raw_requirement or "blank"})
        seen_names[canonical] = seen_names.get(canonical, 0) + 1
        differences = {
            source: details.get(canonical)
            for source, details in catalog.items()
            if details.get(canonical)
        }
        if len(differences) < len(COMMON_EQUITY_SHEETS):
            warning_codes.append("SOURCE_DIFFERENCE")
        instruction = next(
            (details["instructions"] for details in differences.values() if details.get("instructions")), ""
        )
        required_level = _requirement_for(canonical, raw_requirement, order_number)
        maximum = 12 if canonical in {"Sell-Side Reports", "Material Company Press Releases Since Last Earnings Release"} else 5 if canonical in {"Investor Presentations", "Credit Reports"} else 1
        slots.append(
            {
                "order_number": order_number,
                "suborder": suborder,
                "display_name": canonical,
                "normalized_slot_type": _key(canonical).replace(" ", "_"),
                "section_code": section_code,
                "section_name": section_name,
                "required_level": required_level,
                "long_applicable": 1,
                "short_applicable": 1,
                "conditional_rule": "Analyst determines applicability and records a reason." if required_level == "CONDITIONAL" else None,
                "preferred_sources": [preferred_source] if preferred_source else [],
                "fallback_sources": [],
                "instructions": instruction,
                "minimum_documents": 1,
                "maximum_documents": maximum,
                "freshness_rule": None,
                "anchor_rule": None,
                "allowed_document_types": [],
                "expected_output_format": "Source document",
                "auto_search_enabled": 0,
                "manual_upload_allowed": 1,
                "analyst_review_required": 1,
                "default_status": "MANUAL_UPLOAD_REQUIRED",
                "enabled": 1,
                "source_sheet": "MDT",
                "source_row": row,
                "source_coordinates": source_coordinates,
                "raw_import": {
                    "raw_order": raw_order,
                    "raw_name": raw_name,
                    "raw_requirement": raw_requirement,
                    "raw_preferred_source": preferred_source,
                    "source_differences": differences,
                },
                "import_warning": ", ".join(warning_codes) or None,
            }
        )
    for name, count in seen_names.items():
        if count > 1:
            warnings.append({"sheet": "MDT", "code": "DUPLICATE_LABEL", "message": name, "count": count})
    numbered = {slot["order_number"] for slot in slots if slot["suborder"] == 0 and slot["order_number"] is not None}
    gaps = [number for number in range(min(numbered), max(numbered) + 1) if number not in numbered] if numbered else []
    for number in gaps:
        warnings.append({"sheet": "MDT", "code": "MISSING_ORDER_NUMBER", "message": str(number)})
    return {
        "workbook_path": path,
        "workbook_filename": path.name,
        "workbook_sha256": hashlib.sha256(source_bytes).hexdigest(),
        "available_sheets": available,
        "selected_sheets": list(COMMON_EQUITY_SHEETS),
        "raw_rows": raw_rows,
        "slots": slots,
        "warnings": warnings,
        "missing_order_numbers": gaps,
        "unnumbered_supplemental_slots": sum(1 for slot in slots if slot["suborder"]),
        "normalized_slot_count": len(slots),
        "rows_imported": sum(len(rows) for rows in raw_rows.values()),
        "rows_excluded": sum(len(rows) for rows in raw_rows.values()) - len(slots),
        "importer_version": IMPORTER_VERSION,
    }


def import_common_equity_recipe(
    workbook_path: Path | str | None = None,
    *,
    imported_by: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    """Persist a reviewable recipe import; activation remains an explicit later action."""
    report = inspect_common_equity_workbook(workbook_path)
    database.initialize_database(db_path)
    now = database.utc_now_iso()
    import_id = f"RIMP-{_token()}"
    recipe_id = f"RCP-{_token()}"
    with database.get_connection(db_path) as connection:
        version = int(connection.execute(
            "SELECT COALESCE(MAX(version), 0) + 1 FROM package_recipes WHERE recipe_name = ?",
            (RECIPE_NAME,),
        ).fetchone()[0])
        summary = {key: value for key, value in report.items() if key not in {"workbook_path", "raw_rows", "slots"}}
        connection.execute(
            "INSERT INTO recipe_imports VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (import_id, report["workbook_filename"], report["workbook_sha256"], _json(report["available_sheets"]),
             _json(report["selected_sheets"]), IMPORTER_VERSION, now, imported_by, _json(summary)),
        )
        connection.execute(
            """INSERT INTO package_recipes(
                recipe_id, recipe_name, recipe_type, security_type, version, description,
                source_workbook_name, source_workbook_hash, source_sheet, importer_version,
                status, created_at, created_by, notes, import_id
            ) VALUES (?, ?, 'COMMON_EQUITY', 'Common Equity', ?, ?, ?, ?, 'MDT', ?, 'NEEDS_REVIEW', ?, ?, ?, ?)""",
            (recipe_id, RECIPE_NAME, version, "Imported Cutler common-equity package recipe.", report["workbook_filename"],
             report["workbook_sha256"], IMPORTER_VERSION, now, imported_by,
             "Template and Instructions are comparison sources; MDT controls initial ordering.", import_id),
        )
        for sheet_name, rows in report["raw_rows"].items():
            for raw in rows:
                slot = next((item for item in report["slots"] if item["source_sheet"] == sheet_name and item["source_row"] == raw["row_number"]), None)
                connection.execute(
                    "INSERT INTO recipe_import_rows VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (f"RIR-{_token()}", import_id, recipe_id, sheet_name, raw["row_number"], _json(raw["coordinates"]),
                     _json(raw["values"]), _json(slot) if slot else None,
                     _json([slot["import_warning"]]) if slot and slot.get("import_warning") else "[]", int(slot is not None)),
                )
        for slot in report["slots"]:
            connection.execute(
                """INSERT INTO research_slots(
                    slot_id, recipe_id, order_number, suborder, display_name, normalized_slot_type,
                    section_code, section_name, required_level, long_applicable, short_applicable,
                    conditional_rule, preferred_sources_json, fallback_sources_json, instructions,
                    minimum_documents, maximum_documents, freshness_rule, anchor_rule,
                    allowed_document_types_json, expected_output_format, auto_search_enabled,
                    manual_upload_allowed, analyst_review_required, default_status, enabled,
                    source_sheet, source_row, source_coordinates_json, raw_import_json, import_warning, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (f"SLOT-{_token()}", recipe_id, slot["order_number"], slot["suborder"], slot["display_name"],
                 slot["normalized_slot_type"], slot["section_code"], slot["section_name"], slot["required_level"],
                 slot["long_applicable"], slot["short_applicable"], slot["conditional_rule"],
                 _json(slot["preferred_sources"]), _json(slot["fallback_sources"]), slot["instructions"],
                 slot["minimum_documents"], slot["maximum_documents"], slot["freshness_rule"], slot["anchor_rule"],
                 _json(slot["allowed_document_types"]), slot["expected_output_format"], slot["auto_search_enabled"],
                 slot["manual_upload_allowed"], slot["analyst_review_required"], slot["default_status"], slot["enabled"],
                 slot["source_sheet"], slot["source_row"], _json(slot["source_coordinates"]),
                 _json(slot["raw_import"]), slot["import_warning"], now),
            )
        connection.execute(
            "INSERT INTO phase6a_audit_events VALUES (?, NULL, ?, NULL, NULL, 'WORKBOOK_IMPORTED', ?, ?, ?)",
            (f"P6A-{_token()}", recipe_id, _json({"workbook_sha256": report["workbook_sha256"], "slot_count": len(report["slots"]), "warning_count": len(report["warnings"])}), imported_by, now),
        )
        for warning in report["warnings"]:
            connection.execute(
                "INSERT INTO phase6a_audit_events VALUES (?, NULL, ?, NULL, NULL, 'IMPORT_WARNING_CREATED', ?, ?, ?)",
                (f"P6A-{_token()}", recipe_id, _json({"code": warning.get("code"), "sheet": warning.get("sheet"), "row": warning.get("row")}), imported_by, now),
            )
    return {**summary, "import_id": import_id, "recipe_id": recipe_id, "version": version, "status": "NEEDS_REVIEW", "slots": report["slots"]}
