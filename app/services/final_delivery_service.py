from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import uuid
import zipfile
from datetime import date
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app import config
from app.services.final_analysis_service import validate_final_snapshot
from app.services.finalization_service import evaluate_readiness, get_finalization_run
from app.services.package_artifact_service import list_package_artifacts
from app.services.package_recipe_service import board_payload, create_package_from_active_recipe, list_slot_instances
from app.utils import database


SECRET_KEY_RE = re.compile(r"(api[_-]?key|authorization|password|secret|access[_-]?token|environment)", re.I)
SECRET_VALUE_RE = re.compile(r"(?:sk-[A-Za-z0-9_-]{12,}|bearer\s+[A-Za-z0-9._-]{12,})", re.I)
ORDER = {
    "SEC_READER_PDF": 10, "FULL_FILING": 10, "OFFICIAL_DOCUMENT": 30,
    "OFFICIAL_WEB_PAGE": 40, "FILING_SECTION_PDF": 70,
    "LICENSED_UPLOAD": 100, "INTERNAL_UPLOAD": 140,
    "FINAL_RECOMMENDATION": 160, "FINAL_CHECKLIST": 170,
}


def _id(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex.upper()}"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical(value: Any) -> bytes:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")


def _display_date(value: str | None = None) -> str:
    try:
        parsed = date.fromisoformat(str(value or date.today().isoformat())[:10])
    except ValueError:
        parsed = date.today()
    return f"{parsed.month}.{parsed.day}.{str(parsed.year)[-2:]}"


def _safe_name(value: str) -> str:
    clean = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", value)
    clean = re.sub(r"\s+", " ", clean).strip(" .")
    if not clean or clean in {".", ".."}:
        raise ValueError("Package filename is empty or unsafe.")
    if len(clean) > config.FINAL_PACKAGE_MAX_FILENAME_LENGTH:
        suffix = Path(clean).suffix
        clean = clean[: config.FINAL_PACKAGE_MAX_FILENAME_LENGTH - len(suffix)].rstrip(" .") + suffix
    return clean


def _artifact_path(artifact: dict[str, Any]) -> Path:
    return Path(artifact.get("generated_path") or artifact.get("local_path") or "")


def _safe_export(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            key: _safe_export(item)
            for key, item in value.items()
            if key not in {"local_path", "generated_path", "authorization_confirmed"} and not SECRET_KEY_RE.search(str(key))
        }
    if isinstance(value, list):
        return [_safe_export(item) for item in value]
    if isinstance(value, str) and SECRET_VALUE_RE.search(value):
        return "[REDACTED]"
    return value


def build_final_checklist(run_id: str, *, actor: str, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    run = get_finalization_run(run_id, db_path=db_path)
    if not run:
        raise ValueError("Finalization run does not exist.")
    payload = board_payload(run["package_id"], db_path=db_path)
    package = payload["package"]
    artifacts = list_package_artifacts(run["package_id"], include_audit_only=True, db_path=db_path)
    artifacts_by_document: dict[str, list[dict[str, Any]]] = {}
    for artifact in artifacts:
        if artifact.get("source_document_id"):
            artifacts_by_document.setdefault(artifact["source_document_id"], []).append(artifact)
    with database.get_connection(db_path) as connection:
        waivers = {
            row["slot_instance_id"]: dict(row) for row in connection.execute(
                """SELECT * FROM analyst_waivers
                   WHERE package_version_id=? AND status='ACTIVE' AND confirmation_status='CONFIRMED'""",
                (run["package_version_id"],),
            ).fetchall()
        }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Final Package Checklist"
    sheet.freeze_panes = "A7"
    sheet.merge_cells("A1:N1")
    sheet["A1"] = "CUTLER EQUITY RESEARCH - FINAL PACKAGE CHECKLIST"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="171A1F")
    sheet["A2"], sheet["B2"] = "Company", package.get("company_name") or ""
    sheet["D2"], sheet["E2"] = "Ticker", package["ticker"]
    sheet["G2"], sheet["H2"] = "Package Version", run["package_version_id"]
    sheet["A3"], sheet["B3"] = "Research Cutoff", package["research_cutoff_date"]
    sheet["D3"], sheet["E3"] = "Prepared By", actor
    headers = [
        "Order", "Recipe Slot", "Required / Optional", "Availability", "Approved Count",
        "Approved Filenames", "Source", "Document Date", "Highlighted", "Analyst Waiver",
        "Waiver Reason", "Final Inclusion", "Analysis Eligible", "Notes",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(6, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7A1F27")
    readiness = evaluate_readiness(run["package_id"], package_version_id=run["package_version_id"], db_path=db_path)
    states = {row["slot_instance_id"]: row["status"] for row in readiness.slots}
    for row_number, slot in enumerate(payload["slots"], 7):
        assignments = [row for row in payload["assignments_by_slot"].get(slot["package_slot_instance_id"], [])
                       if row["assignment_status"] == "APPROVED" and row["selected_for_package"]]
        final_artifacts = [artifact for assignment in assignments for artifact in artifacts_by_document.get(assignment["document_id"], [])
                           if artifact["artifact_status"] == "CURRENT" and artifact["working_package_inclusion"]]
        waiver = waivers.get(slot["package_slot_instance_id"])
        values = [
            f"{slot.get('order_number') or ''}{'.' + str(slot['suborder']) if slot.get('suborder') else ''}",
            slot["display_name_snapshot"], slot["requirement_snapshot"], states.get(slot["package_slot_instance_id"], "UNKNOWN"),
            len(assignments), "; ".join(sorted({artifact["display_filename"] for artifact in final_artifacts})),
            "; ".join(sorted({str(row.get("source_name") or "") for row in assignments if row.get("source_name")})),
            "; ".join(sorted({str(row.get("document_date") or row.get("publication_date") or "") for row in assignments})),
            "Yes" if any(row.get("highlighted_research") for row in assignments) else "No",
            "Yes" if waiver else "No", waiver.get("reason") if waiver else "",
            "Included" if final_artifacts else "Unavailable / Not Applicable" if states.get(slot["package_slot_instance_id"]) in {"ACKNOWLEDGED_UNAVAILABLE", "NOT_APPLICABLE"} else "Not Included",
            "Yes" if any(row.get("analysis_eligible") for row in final_artifacts) else "No",
            slot.get("analyst_notes") or "",
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
    widths = [10, 34, 20, 26, 15, 48, 24, 16, 12, 15, 40, 20, 18, 42]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    filename = _safe_name(f"{package['ticker']} Equity Research Checklist {_display_date(package.get('compilation_date'))}.xlsx")
    path = config.PACKAGE_DIR / run["package_id"] / "phase6c" / "final" / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    load_workbook(path, read_only=True, data_only=False).close()
    digest = _sha256(path)
    artifact_id = "ART-" + hashlib.sha256(f"{run['package_id']}|FINAL_CHECKLIST|{run['package_version_id']}".encode()).hexdigest()[:20].upper()
    now = database.utc_now_iso()
    with database.get_connection(db_path) as connection:
        connection.execute(
            """UPDATE package_artifacts SET artifact_status='SUPERSEDED', working_package_inclusion=0, superseded_at=?
               WHERE package_id=? AND artifact_type IN ('CHECKLIST','FINAL_CHECKLIST') AND artifact_status='CURRENT' AND artifact_id!=?""",
            (now, run["package_id"], artifact_id),
        )
        connection.execute(
            """INSERT INTO package_artifacts(
               artifact_id, package_id, artifact_type, display_filename, purpose_label,
               working_package_inclusion, audit_package_inclusion, analysis_eligible,
               conversion_status, artifact_status, created_at, generated_path, generated_sha256,
               generated_size_bytes, qa_status, qa_result_json, source_role, package_version_id
               ) VALUES (?, ?, 'FINAL_CHECKLIST', ?, 'Final Package Checklist', 1, 1, 0,
               'FINAL_CHECKLIST_READY', 'CURRENT', ?, ?, ?, ?, 'PASSED', ?, 'ANALYST_NOTE', ?)""",
            (artifact_id, run["package_id"], filename, now, str(path), digest, path.stat().st_size,
             json.dumps({"passed": True, "workbook_opens": True}, sort_keys=True), run["package_version_id"]),
        )
    return {"artifact_id": artifact_id, "path": str(path), "sha256": digest, "filename": filename}


def _working_artifacts(run: dict[str, Any], db_path: Path | str) -> list[dict[str, Any]]:
    rows = [row for row in list_package_artifacts(run["package_id"], db_path=db_path)
            if row["working_package_inclusion"] and row["artifact_status"] == "CURRENT"]
    ready = []
    names: set[str] = set()
    for row in rows:
        path = _artifact_path(row)
        if not path.is_file():
            raise ValueError(f"Working-package artifact is missing: {row['display_filename']}")
        name = _safe_name(row["display_filename"])
        if name.casefold() in names:
            raise ValueError(f"Duplicate working-package filename: {name}")
        if row["artifact_type"] in {"SEC_READER_PDF", "FILING_SECTION_PDF", "FINAL_RECOMMENDATION"} and row.get("qa_status") != "PASSED":
            raise ValueError(f"Generated PDF did not pass QA: {name}")
        names.add(name.casefold())
        ready.append({**row, "display_filename": name, "resolved_path": str(path)})
    ready.sort(key=lambda row: (ORDER.get(row["artifact_type"], 90), row.get("order_number") or 999, row["display_filename"].casefold()))
    return ready


def build_manifest(run_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    run = get_finalization_run(run_id, db_path=db_path)
    if not run:
        raise ValueError("Finalization run does not exist.")
    package = database.get_package_by_package_id(run["package_id"], db_path=db_path) or {}
    payload = board_payload(run["package_id"], db_path=db_path)
    artifacts = _working_artifacts(run, db_path)
    with database.get_connection(db_path) as connection:
        snapshot = connection.execute(
            "SELECT * FROM final_analysis_snapshots WHERE package_version_id=? AND status='READY' ORDER BY created_at DESC LIMIT 1",
            (run["package_version_id"],),
        ).fetchone()
        approval = connection.execute(
            "SELECT * FROM final_recommendation_approvals WHERE package_version_id=? ORDER BY created_at DESC LIMIT 1",
            (run["package_version_id"],),
        ).fetchone()
    if not snapshot or not approval or approval["status"] != "APPROVED":
        raise ValueError("Final snapshot and approved recommendation are required before manifest creation.")
    root = f"{package['ticker']} Equity Research Package"
    files = []
    for artifact in artifacts:
        path = Path(artifact["resolved_path"])
        files.append({
            "relative_archive_path": f"{root}/{artifact['display_filename']}",
            "artifact_id": artifact["artifact_id"], "artifact_type": artifact["artifact_type"],
            "checklist_slot": artifact.get("checklist_item"), "source_role": artifact.get("source_role"),
            "source_document_id": artifact.get("source_document_id"), "source_hash": artifact.get("sha256_hash"),
            "generated_artifact_hash": artifact.get("generated_sha256") or _sha256(path),
            "file_size": path.stat().st_size, "document_date": artifact.get("document_date") or artifact.get("publication_date"),
            "analysis_eligibility": bool(artifact.get("analysis_eligible")),
            "highlighted_status": False, "conversion_status": artifact.get("conversion_status"),
        })
    manifest = {
        "manifest_version": "6C.0", "package_id": run["package_id"],
        "package_version": run["package_version_id"], "company": package.get("company_name"),
        "ticker": package.get("ticker"), "cik": package.get("cik"),
        "recipe_name": payload["recipe"].get("recipe_name"), "recipe_version": payload["recipe"].get("version"),
        "research_cutoff": package.get("research_cutoff_date"),
        "earnings_cycle": payload.get("summary", {}).get("earnings_cycle"),
        "final_snapshot_id": snapshot["final_snapshot_id"],
        "final_recommendation": approval["analyst_recommendation"], "confidence": approval["ai_confidence"],
        "analyst_approval": {"approved_by": approval["approved_by"], "approved_at": approval["approved_at"]},
        "lock_timestamp": None, "artifact_count": len(files), "working_package_files": files,
    }
    raw = _canonical(manifest)
    if SECRET_VALUE_RE.search(raw.decode("utf-8")):
        raise ValueError("Manifest secret scan failed.")
    digest = hashlib.sha256(raw).hexdigest()
    path = config.PACKAGE_DIR / run["package_id"] / "phase6c" / "final" / "package_manifest.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(raw)
    with database.get_connection(db_path) as connection:
        connection.execute(
            """INSERT INTO final_package_manifests VALUES (?, ?, ?, 'FINAL', ?, ?, ?, ?)
               ON CONFLICT(package_version_id, manifest_type) DO UPDATE SET
                 manifest_json=excluded.manifest_json, manifest_sha256=excluded.manifest_sha256,
                 local_path=excluded.local_path, created_at=excluded.created_at""",
            (_id("MAN"), run["package_id"], run["package_version_id"], raw.decode("utf-8"), digest, str(path), database.utc_now_iso()),
        )
    return {"manifest": manifest, "sha256": digest, "path": str(path), "artifacts": artifacts}


def _zip_entry_from_file(archive: zipfile.ZipFile, archive_path: str, source: Path) -> None:
    pure = PurePosixPath(archive_path)
    if pure.is_absolute() or ".." in pure.parts or "" in pure.parts:
        raise ValueError("Unsafe ZIP archive path.")
    info = zipfile.ZipInfo(str(pure), date_time=(1980, 1, 1, 0, 0, 0))
    info.compress_type = zipfile.ZIP_DEFLATED
    info.external_attr = 0o100644 << 16
    with source.open("rb") as source_stream, archive.open(info, "w") as target:
        for chunk in iter(lambda: source_stream.read(1024 * 1024), b""):
            target.write(chunk)


def _zip_entry_bytes(archive: zipfile.ZipFile, archive_path: str, content: bytes) -> None:
    info = zipfile.ZipInfo(archive_path, date_time=(1980, 1, 1, 0, 0, 0))
    info.compress_type = zipfile.ZIP_DEFLATED
    info.external_attr = 0o100644 << 16
    archive.writestr(info, content)


def _qa_zip(path: Path, *, expected_paths: Iterable[str]) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        bad = archive.testzip()
    safe = all(not PurePosixPath(name).is_absolute() and ".." not in PurePosixPath(name).parts for name in names)
    checks = {
        "opens": True, "entry_count": len(names), "no_duplicate_paths": len(names) == len(set(names)),
        "safe_paths": safe, "crc_passed": bad is None, "expected_paths_present": set(expected_paths) <= set(names),
    }
    checks["passed"] = all(checks.values())
    return checks


def build_working_zip(run_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    built = build_manifest(run_id, db_path=db_path)
    run = get_finalization_run(run_id, db_path=db_path) or {}
    package = database.get_package_by_package_id(run["package_id"], db_path=db_path) or {}
    filename = _safe_name(f"{package['ticker']} Equity Research Package {_display_date(package.get('compilation_date'))}.zip")
    path = config.PACKAGE_DIR / run["package_id"] / "phase6c" / "final" / filename
    expected = []
    with zipfile.ZipFile(path, "w", allowZip64=True) as archive:
        for artifact, item in zip(built["artifacts"], built["manifest"]["working_package_files"], strict=True):
            expected.append(item["relative_archive_path"])
            _zip_entry_from_file(archive, item["relative_archive_path"], Path(artifact["resolved_path"]))
    qa = _qa_zip(path, expected_paths=expected)
    result = {"path": str(path), "sha256": _sha256(path), "file_size_bytes": path.stat().st_size, "qa": qa}
    _store_zip(run, "WORKING", result, db_path)
    return result


def _query_rows(connection: Any, sql: str, params: tuple[Any, ...]) -> list[dict[str, Any]]:
    return [_safe_export(dict(row)) for row in connection.execute(sql, params).fetchall()]


def build_audit_zip(run_id: str, *, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    built = build_manifest(run_id, db_path=db_path)
    run = get_finalization_run(run_id, db_path=db_path) or {}
    package = database.get_package_by_package_id(run["package_id"], db_path=db_path) or {}
    filename = _safe_name(f"{package['ticker']} Equity Research Audit {_display_date(package.get('compilation_date'))}.zip")
    path = config.PACKAGE_DIR / run["package_id"] / "phase6c" / "final" / filename
    audit_records: dict[str, Any] = {}
    with database.get_connection(db_path) as connection:
        for name, sql in {
            "recipe_snapshot": "SELECT recipe_snapshot_json FROM package_recipe_instances WHERE package_id=?",
            "analysis_snapshot": "SELECT * FROM final_analysis_snapshots WHERE package_version_id=?",
            "deterministic_facts": "SELECT * FROM normalized_financial_facts WHERE package_version_id=?",
            "conflicts": "SELECT * FROM financial_fact_conflicts WHERE package_version_id=?",
            "numeric_claims": "SELECT nc.* FROM numeric_claims nc JOIN analysis_corpus_snapshots acs ON acs.snapshot_id=nc.snapshot_id WHERE acs.package_id=?",
            "analyst_waivers": "SELECT * FROM analyst_waivers WHERE package_version_id=?",
            "analyst_overrides": "SELECT aro.* FROM analyst_rating_overrides aro JOIN final_recommendation_approvals fra ON fra.approval_id=aro.approval_id WHERE fra.package_version_id=?",
            "report_qa": "SELECT * FROM final_recommendation_approvals WHERE package_version_id=?",
            "search_audit": "SELECT candidate_id, source_provider, source_route, title, canonical_url, candidate_status, rejection_reason_code, created_at FROM discovered_candidates WHERE package_id=?",
            "candidate_decisions": "SELECT * FROM candidate_decisions WHERE package_id=?",
            "OpenAI_usage": """SELECT oul.model, oul.endpoint, oul.input_tokens, oul.cached_input_tokens,
                                oul.output_tokens, oul.total_tokens, oul.estimated_cost_usd, oul.output_status, oul.created_at
                                FROM openai_usage_ledger oul JOIN analysis_runs ar ON ar.analysis_run_id=oul.analysis_run_id
                                WHERE ar.package_id=?""",
            "Brave_usage": """SELECT du.* FROM discovery_usage du
                               JOIN package_discovery_runs pdr ON pdr.discovery_run_id=du.discovery_run_id
                               WHERE pdr.package_id=?""",
        }.items():
            parameter = run["package_version_id"] if name in {"analysis_snapshot", "deterministic_facts", "conflicts", "analyst_waivers", "analyst_overrides", "report_qa"} else run["package_id"]
            audit_records[name] = _query_rows(connection, sql, (parameter,))
        schema = connection.execute("SELECT schema_value FROM schema_metadata WHERE schema_key='database_schema_version'").fetchone()[0]
    expected = ["Audit/package_manifest.json", "Audit/migration_and_version_info/schema.json"]
    with zipfile.ZipFile(path, "w", allowZip64=True) as archive:
        _zip_entry_bytes(archive, "Audit/package_manifest.json", _canonical(built["manifest"]))
        _zip_entry_bytes(archive, "Audit/migration_and_version_info/schema.json", _canonical({"database_schema_version": schema, "phase": "6C"}))
        for name, rows in sorted(audit_records.items()):
            entry = f"Audit/{name}/{name}.json"
            expected.append(entry)
            content = _canonical(rows)
            if SECRET_VALUE_RE.search(content.decode("utf-8")):
                raise ValueError(f"Audit secret scan failed for {name}.")
            _zip_entry_bytes(archive, entry, content)
        for artifact in list_package_artifacts(run["package_id"], include_audit_only=True, db_path=db_path):
            source = Path(artifact.get("local_path") or "")
            if not source.is_file() or not artifact.get("source_document_id"):
                continue
            role = str(artifact.get("source_role") or "")
            include_bytes = role in {"SEC_PRIMARY", "COMPANY_PRIMARY"} or config.LICENSED_AUDIT_BYTES_ENABLED
            if include_bytes:
                folder = "original_sec_html" if source.suffix.casefold() in {".html", ".htm"} and role == "SEC_PRIMARY" else "source_documents"
                entry = f"Audit/{folder}/{artifact['source_document_id']}-{_safe_name(source.name)}"
                if entry not in expected:
                    expected.append(entry)
                    _zip_entry_from_file(archive, entry, source)
            else:
                entry = f"Audit/document_hashes/{artifact['source_document_id']}.json"
                if entry not in expected:
                    expected.append(entry)
                    _zip_entry_bytes(archive, entry, _canonical({"document_id": artifact["source_document_id"], "sha256": artifact.get("sha256_hash"), "licensed_bytes_included": False}))
    qa = _qa_zip(path, expected_paths=expected)
    result = {"path": str(path), "sha256": _sha256(path), "file_size_bytes": path.stat().st_size, "qa": qa}
    _store_zip(run, "AUDIT", result, db_path)
    return result


def _store_zip(run: dict[str, Any], zip_type: str, result: dict[str, Any], db_path: Path | str) -> None:
    with database.get_connection(db_path) as connection:
        connection.execute(
            """INSERT INTO final_zip_outputs VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(package_version_id, zip_type) DO UPDATE SET local_path=excluded.local_path,
               sha256=excluded.sha256, file_size_bytes=excluded.file_size_bytes,
               qa_status=excluded.qa_status, qa_result_json=excluded.qa_result_json, created_at=excluded.created_at""",
            (_id("ZIP"), run["package_id"], run["package_version_id"], zip_type, result["path"], result["sha256"],
             result["file_size_bytes"], "PASSED" if result["qa"]["passed"] else "FAILED",
             json.dumps(result["qa"], sort_keys=True), database.utc_now_iso()),
        )


def run_final_qa(run_id: str, *, actor: str, db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    run = get_finalization_run(run_id, db_path=db_path)
    if not run:
        raise ValueError("Finalization run does not exist.")
    readiness = evaluate_readiness(run["package_id"], package_version_id=run["package_version_id"], db_path=db_path)
    with database.get_connection(db_path) as connection:
        snapshot = connection.execute("SELECT * FROM final_analysis_snapshots WHERE package_version_id=? ORDER BY created_at DESC LIMIT 1", (run["package_version_id"],)).fetchone()
        approval = connection.execute("SELECT * FROM final_recommendation_approvals WHERE package_version_id=? ORDER BY created_at DESC LIMIT 1", (run["package_version_id"],)).fetchone()
        manifest = connection.execute("SELECT * FROM final_package_manifests WHERE package_version_id=? AND manifest_type='FINAL'", (run["package_version_id"],)).fetchone()
        zips = {row["zip_type"]: dict(row) for row in connection.execute("SELECT * FROM final_zip_outputs WHERE package_version_id=?", (run["package_version_id"],)).fetchall()}
        checklist = connection.execute("SELECT * FROM package_artifacts WHERE package_version_id=? AND artifact_type='FINAL_CHECKLIST' AND artifact_status='CURRENT'", (run["package_version_id"],)).fetchone()
    checks = {
        "readiness": readiness.ready,
        "snapshot": bool(snapshot) and validate_final_snapshot(snapshot["final_snapshot_id"], db_path=db_path)["status"] == "PASSED",
        "recommendation": bool(approval) and approval["status"] == "APPROVED" and approval["qa_status"] == "PASSED",
        "checklist": bool(checklist) and Path(checklist["generated_path"]).is_file() and _sha256(Path(checklist["generated_path"])) == checklist["generated_sha256"],
        "manifest": bool(manifest) and hashlib.sha256(manifest["manifest_json"].encode("utf-8")).hexdigest() == manifest["manifest_sha256"],
        "working_zip": zips.get("WORKING", {}).get("qa_status") == "PASSED" and Path(zips.get("WORKING", {}).get("local_path", "")).is_file(),
        "audit_zip": zips.get("AUDIT", {}).get("qa_status") == "PASSED" and Path(zips.get("AUDIT", {}).get("local_path", "")).is_file(),
    }
    status = "PASSED" if all(checks.values()) else "FAILED"
    qa_id = _id("FQA")
    with database.get_connection(db_path) as connection:
        connection.execute("INSERT INTO final_qa_results VALUES (?, ?, ?, ?, ?, ?, ?)",
                           (qa_id, run["package_id"], run["package_version_id"], status, json.dumps(checks, sort_keys=True), database.utc_now_iso(), actor))
    return {"final_qa_id": qa_id, "status": status, "checks": checks}


def lock_final_package(
    run_id: str,
    *,
    actor: str,
    analyst_confirmed: bool,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    if not analyst_confirmed:
        raise ValueError("Analyst confirmation is required to lock the final package.")
    run = get_finalization_run(run_id, db_path=db_path)
    if not run:
        raise ValueError("Finalization run does not exist.")
    with database.get_connection(db_path) as connection:
        qa = connection.execute("SELECT * FROM final_qa_results WHERE package_version_id=? ORDER BY created_at DESC LIMIT 1", (run["package_version_id"],)).fetchone()
        manifest = connection.execute("SELECT * FROM final_package_manifests WHERE package_version_id=? AND manifest_type='FINAL'", (run["package_version_id"],)).fetchone()
        snapshot = connection.execute("SELECT * FROM final_analysis_snapshots WHERE package_version_id=? ORDER BY created_at DESC LIMIT 1", (run["package_version_id"],)).fetchone()
        zips = {row["zip_type"]: row for row in connection.execute("SELECT * FROM final_zip_outputs WHERE package_version_id=?", (run["package_version_id"],)).fetchall()}
        if not qa or qa["status"] != "PASSED" or not manifest or not snapshot or set(zips) != {"WORKING", "AUDIT"}:
            raise ValueError("Final QA, manifest, snapshot, and both ZIPs must pass before locking.")
        payload = {
            "package_id": run["package_id"], "package_version_id": run["package_version_id"],
            "final_snapshot_id": snapshot["final_snapshot_id"], "final_snapshot_hash": snapshot["snapshot_hash"],
            "manifest_sha256": manifest["manifest_sha256"], "working_zip_sha256": zips["WORKING"]["sha256"],
            "audit_zip_sha256": zips["AUDIT"]["sha256"], "final_qa_id": qa["final_qa_id"],
        }
        lock_id = _id("LOCK")
        now = database.utc_now_iso()
        connection.execute("UPDATE package_versions SET status='LOCKED', locked_at=? WHERE version_id=?", (now, run["package_version_id"]))
        connection.execute("UPDATE packages SET status=?, updated_at=? WHERE package_id=?", (config.STATUS_PACKAGE_LOCKED, now, run["package_id"]))
        connection.execute(
            "INSERT INTO final_package_locks VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (lock_id, run["package_id"], run["package_version_id"], snapshot["final_snapshot_id"], manifest["manifest_sha256"],
             zips["WORKING"]["sha256"], zips["AUDIT"]["sha256"], actor, now, json.dumps(payload, sort_keys=True)),
        )
        connection.execute("UPDATE finalization_runs SET status='LOCKED', current_stage='LOCKED', updated_at=?, completed_at=? WHERE finalization_run_id=?", (now, now, run_id))
        row = connection.execute("SELECT * FROM final_package_locks WHERE lock_id=?", (lock_id,)).fetchone()
    return dict(row)


def mark_delivered(run_id: str, *, actor: str, note: str = "", db_path: Path | str = config.DATABASE_PATH) -> dict[str, Any]:
    run = get_finalization_run(run_id, db_path=db_path)
    if not run:
        raise ValueError("Finalization run does not exist.")
    with database.get_connection(db_path) as connection:
        lock = connection.execute("SELECT * FROM final_package_locks WHERE package_version_id=?", (run["package_version_id"],)).fetchone()
        zips = {row["zip_type"]: row for row in connection.execute("SELECT * FROM final_zip_outputs WHERE package_version_id=?", (run["package_version_id"],)).fetchall()}
        if not lock:
            raise ValueError("Only a locked package can be marked delivered.")
        delivery_id, now = _id("DLV"), database.utc_now_iso()
        connection.execute("INSERT INTO delivery_records VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                           (delivery_id, run["package_id"], run["package_version_id"], actor, now,
                            zips["WORKING"]["sha256"], zips["AUDIT"]["sha256"], note.strip() or None))
        connection.execute("UPDATE finalization_runs SET status='DELIVERED', current_stage='DELIVERED', updated_at=? WHERE finalization_run_id=?", (now, run_id))
        row = connection.execute("SELECT * FROM delivery_records WHERE delivery_id=?", (delivery_id,)).fetchone()
    return dict(row)


def create_new_package_version(
    locked_package_id: str,
    *,
    actor: str,
    db_path: Path | str = config.DATABASE_PATH,
) -> dict[str, Any]:
    """Clone locked configuration, source records, and assignments into a distinct draft package."""
    source = database.get_package_by_package_id(locked_package_id, db_path=db_path)
    if not source:
        raise ValueError("Locked package does not exist.")
    with database.get_connection(db_path) as connection:
        lock = connection.execute("SELECT * FROM final_package_locks WHERE package_id=?", (locked_package_id,)).fetchone()
    if not lock:
        raise ValueError("Create New Package Version is available only for a final locked package.")
    try:
        cutoff = date.fromisoformat(str(source["research_cutoff_date"])[:10])
    except ValueError:
        cutoff = date.today()
    cloned = create_package_from_active_recipe(
        source, research_cutoff=cutoff,
        compilation_date=date.today(), compiled_by=actor, created_by=actor, db_path=db_path,
    )
    source_slots = {row["slot_id"]: row for row in list_slot_instances(locked_package_id, db_path=db_path)}
    target_slots = {row["slot_id"]: row for row in list_slot_instances(cloned["package_id"], db_path=db_path)}
    now = database.utc_now_iso()
    document_map: dict[str, str] = {}
    assignment_count = 0
    with database.get_connection(db_path) as connection:
        connection.execute("UPDATE packages SET source_legacy_package_id=? WHERE package_id=?", (locked_package_id, cloned["package_id"]))
        columns = [row["name"] for row in connection.execute("PRAGMA table_info(documents)").fetchall() if row["name"] != "id"]
        for document in connection.execute("SELECT * FROM documents WHERE package_id=?", (locked_package_id,)).fetchall():
            copied = {column: document[column] for column in columns}
            new_id = database.generate_document_id("DOC-VERSION")
            copied.update({"document_id": new_id, "package_id": cloned["package_id"], "ticker": cloned["ticker"],
                           "archive_origin_document_id": document["document_id"], "created_at": now, "updated_at": now})
            connection.execute(f"INSERT INTO documents({', '.join(columns)}) VALUES ({', '.join('?' for _ in columns)})",
                               tuple(copied[column] for column in columns))
            document_map[document["document_id"]] = new_id
        assignments = connection.execute("SELECT * FROM slot_document_assignments WHERE package_id=?", (locked_package_id,)).fetchall()
        for assignment in assignments:
            source_slot = connection.execute("SELECT slot_id FROM package_slot_instances WHERE package_slot_instance_id=?", (assignment["package_slot_instance_id"],)).fetchone()
            target_slot = target_slots.get(source_slot[0]) if source_slot else None
            new_document_id = document_map.get(assignment["document_id"])
            if not target_slot or not new_document_id:
                continue
            connection.execute(
                """INSERT INTO slot_document_assignments(
                   assignment_id, package_slot_instance_id, package_id, document_id, assignment_source,
                   suggested_slot_id, final_slot_id, suggestion_confidence, suggestion_reason, matched_tokens_json,
                   assignment_status, selected_for_package, highlighted_research, display_order, analyst_notes,
                   assigned_at, assigned_by, approved_at, approved_by, replaced_assignment_id
                   ) VALUES (?, ?, ?, ?, 'COPIED_FROM_LOCKED_VERSION', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL)""",
                (database.generate_document_id("ASG-VERSION"), target_slot["package_slot_instance_id"], cloned["package_id"],
                 new_document_id, assignment["suggested_slot_id"], assignment["final_slot_id"], assignment["suggestion_confidence"],
                 assignment["suggestion_reason"], assignment["matched_tokens_json"], assignment["assignment_status"],
                 assignment["selected_for_package"], assignment["highlighted_research"], assignment["display_order"],
                 assignment["analyst_notes"], now, actor, now if assignment["assignment_status"] == "APPROVED" else None,
                 actor if assignment["assignment_status"] == "APPROVED" else None),
            )
            assignment_count += 1
        for slot_id, source_slot in source_slots.items():
            target_slot = target_slots.get(slot_id)
            if target_slot:
                connection.execute(
                    """UPDATE package_slot_instances SET applicability_status=?, completion_status=?,
                       analyst_acknowledged=?, analyst_notes=?, updated_at=? WHERE package_slot_instance_id=?""",
                    (source_slot["applicability_status"], source_slot["completion_status"], source_slot["analyst_acknowledged"],
                     source_slot["analyst_notes"], now, target_slot["package_slot_instance_id"]),
                )
        connection.execute(
            """INSERT INTO phase6a_audit_events(event_id, event_type, actor, package_id, event_details_json, created_at)
               VALUES (?, 'NEW_PACKAGE_VERSION_CREATED', ?, ?, ?, ?)""",
            (_id("AUD"), actor, cloned["package_id"], json.dumps({"source_locked_package_id": locked_package_id,
             "source_lock_id": lock["lock_id"], "copied_documents": len(document_map), "copied_assignments": assignment_count}, sort_keys=True), now),
        )
    from app.services.package_recipe_service import recalculate_completion
    from app.services.package_artifact_service import sync_package_artifacts

    recalculate_completion(cloned["package_id"], actor=actor, db_path=db_path)
    sync_package_artifacts(cloned["package_id"], db_path=db_path)
    return database.get_package_by_package_id(cloned["package_id"], db_path=db_path) or cloned
