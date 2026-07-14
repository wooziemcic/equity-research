from __future__ import annotations

import csv
import json
import zipfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from app import config
from app.services.checklist_service import ensure_package_checklist, set_override
from app.services.package_builder import (
    build_package_version,
    compare_versions,
    lock_version,
    validate_package_readiness,
    verify_snapshot,
)
from app.services.package_service import PackageInput, create_package
from app.services.upload_service import UploadCandidate, store_uploaded_files
from app.utils import database


@pytest.fixture(autouse=True)
def phase4_config(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(config, "DOWNLOAD_DIR", tmp_path / "downloaded")
    monkeypatch.setattr(config, "PACKAGE_DIR", tmp_path / "packages")
    monkeypatch.setattr(config, "MAX_UPLOAD_FILE_MB", 5)
    monkeypatch.setattr(config, "MAX_UPLOAD_BATCH_MB", 20)


@pytest.fixture()
def temp_db(tmp_path: Path) -> Path:
    db_path = tmp_path / "phase4.db"
    database.initialize_database(db_path)
    return db_path


def prepared_package(temp_db: Path, ticker: str = "QXO") -> dict:
    package = create_package(
        PackageInput(ticker, "Common Equity", date(2026, 7, 13), 3, ""),
        db_path=temp_db,
    )
    package = database.update_package_company_metadata(
        package["package_id"],
        {
            "ticker": ticker,
            "company_name": f"{ticker} Inc.",
            "cik": "0001234567",
            "exchange": "NYSE",
            "sic": "7370",
            "industry_description": "Services",
            "fiscal_year_end": "1231",
            "sec_company_url": "https://www.sec.gov/edgar/browse/?CIK=0001234567",
            "resolution_status": "RESOLVED",
            "resolution_source": "test",
            "resolution_timestamp": "2026-07-13T00:00:00+00:00",
        },
        db_path=temp_db,
    )
    store_uploaded_files(
        package,
        [
            UploadCandidate("QXO 10-K.pdf", b"%PDF-1.4 annual"),
            UploadCandidate("Bloomberg DES.pdf", b"%PDF-1.4 des"),
        ],
        source_type="bloomberg",
        authorization_confirmed=True,
        metadata_by_name={
            "QXO 10-K.pdf": {"final_category_code": "annual_filing", "title": "Annual filing"},
            "Bloomberg DES.pdf": {"final_category_code": "bloomberg_des", "title": "Bloomberg DES"},
        },
        db_path=temp_db,
    )
    ensure_package_checklist(package, db_path=temp_db)
    package = database.update_package_review_acknowledgement(
        package["package_id"],
        checklist_reviewed=True,
        reviewed_by="analyst",
        review_note="Reviewed for package build.",
        missing_core_acknowledged=True,
        stale_documents_acknowledged=True,
        needs_review_acknowledged=True,
        db_path=temp_db,
    )
    return package


def test_readiness_validation_states(temp_db: Path) -> None:
    missing = validate_package_readiness(None, db_path=temp_db)
    assert missing.status == config.READINESS_NOT_READY
    package = create_package(PackageInput("ABC", "Common Equity", date.today(), 3, ""), db_path=temp_db)
    no_docs = validate_package_readiness(package, db_path=temp_db)
    assert "At least one document is required." in no_docs.errors
    package = prepared_package(temp_db)
    ready = validate_package_readiness(package, db_path=temp_db)
    assert ready.status in {config.READINESS_READY, config.READINESS_READY_WITH_WARNINGS}
    unreviewed = database.update_package_review_acknowledgement(
        package["package_id"],
        checklist_reviewed=False,
        reviewed_by="analyst",
        review_note="",
        missing_core_acknowledged=True,
        stale_documents_acknowledged=True,
        needs_review_acknowledged=True,
        db_path=temp_db,
    )
    assert validate_package_readiness(unreviewed, db_path=temp_db).status == config.READINESS_NOT_READY


def test_readiness_blocks_bad_documents(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    doc = database.list_documents_by_package(package["package_id"], db_path=temp_db)[0]
    database.update_document_status(doc["document_id"], config.DOCUMENT_STATUS_FAILED, error_message="bad", db_path=temp_db)
    assert any("Failed document" in error for error in validate_package_readiness(package, db_path=temp_db).errors)
    database.update_document_status(doc["document_id"], config.DOCUMENT_STATUS_DOWNLOADED, db_path=temp_db)
    database.update_document_metadata(doc["document_id"], {"category": "", "final_category_code": ""}, db_path=temp_db)
    assert any("missing a category" in error for error in validate_package_readiness(package, db_path=temp_db).errors)


def test_readiness_blocks_missing_hash_missing_file_and_outside_path(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    doc = database.list_documents_by_package(package["package_id"], db_path=temp_db)[0]
    database.update_document_metadata(doc["document_id"], {"category": "Annual Filing"}, db_path=temp_db)
    with database.get_connection(temp_db) as connection:
        connection.execute("UPDATE documents SET sha256_hash = NULL WHERE document_id = ?", (doc["document_id"],))
    assert any("missing SHA-256" in error for error in validate_package_readiness(package, db_path=temp_db).errors)
    with database.get_connection(temp_db) as connection:
        connection.execute("UPDATE documents SET sha256_hash = ?, local_path = ? WHERE document_id = ?", ("abc", "C:/outside/file.pdf", doc["document_id"]))
    assert any("invalid managed path" in error for error in validate_package_readiness(package, db_path=temp_db).errors)
    with database.get_connection(temp_db) as connection:
        connection.execute("UPDATE documents SET local_path = ? WHERE document_id = ?", (str(config.DOWNLOAD_DIR / "missing.pdf"), doc["document_id"]))
    assert any("file is missing" in error for error in validate_package_readiness(package, db_path=temp_db).errors)


def test_stale_and_needs_review_require_acknowledgement(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    set_override(package["package_id"], "latest_annual", config.CHECKLIST_STATUS_STALE, "Old", db_path=temp_db)
    package = database.update_package_review_acknowledgement(
        package["package_id"], checklist_reviewed=True, reviewed_by="analyst", review_note="", missing_core_acknowledged=True, stale_documents_acknowledged=False, needs_review_acknowledged=True, db_path=temp_db
    )
    assert any("Stale" in error for error in validate_package_readiness(package, db_path=temp_db).errors)
    set_override(package["package_id"], "latest_annual", config.CHECKLIST_STATUS_NEEDS_REVIEW, "Review", db_path=temp_db)
    package = database.update_package_review_acknowledgement(
        package["package_id"], checklist_reviewed=True, reviewed_by="analyst", review_note="", missing_core_acknowledged=True, stale_documents_acknowledged=True, needs_review_acknowledged=False, db_path=temp_db
    )
    assert any("Needs-review" in error for error in validate_package_readiness(package, db_path=temp_db).errors)


def test_build_version_structure_manifest_inventory_zip_and_lock(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    version = build_package_version(package, notes="first build", db_path=temp_db)
    assert version["version_id"].startswith("PV-")
    assert version["display_version"].endswith("V001")
    assert version["status"] == config.VERSION_STATUS_BUILT
    root = Path(version["manifest_path"]).parents[1]
    assert (root / "00_Package_Manifest" / "package_manifest.json").exists()
    assert (root / "01_SEC_Filings").exists()
    assert (root / "04_Bloomberg").exists()
    manifest = json.loads(Path(version["manifest_path"]).read_text(encoding="utf-8"))
    assert manifest["package_id"] == package["package_id"]
    assert manifest["document_counts"]["licensed"] == 2
    assert manifest["documents"] == sorted(manifest["documents"], key=lambda doc: doc["document_id"])
    assert version["manifest_sha256"]
    with Path(version["inventory_path"]).open(encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 2
    workbook = load_workbook(root / "00_Package_Manifest" / "document_inventory.xlsx", data_only=False)
    sheet = workbook.active
    assert sheet.freeze_panes == "A2"
    assert all(not (isinstance(cell.value, str) and cell.value.startswith("=")) for row in sheet.iter_rows() for cell in row)
    with zipfile.ZipFile(version["zip_path"]) as archive:
        names = archive.namelist()
    assert "00_Package_Manifest/package_manifest.json" in names
    assert all(not Path(name).is_absolute() for name in names)
    assert not any(name.endswith(".db") or ".env" in name for name in names)
    locked = lock_version(version["version_id"], db_path=temp_db)
    assert locked["status"] == config.VERSION_STATUS_LOCKED
    assert locked["locked_at"]


def test_unchanged_second_build_reuses_content_addressed_snapshot(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    first = build_package_version(package, db_path=temp_db)
    second = build_package_version(package, db_path=temp_db)
    assert first["display_version"].endswith("V001")
    assert second["version_id"] == first["version_id"]
    assert second["zip_path"] == first["zip_path"]
    assert Path(first["zip_path"]).exists()


def test_failed_validation_does_not_create_version(temp_db: Path) -> None:
    package = create_package(PackageInput("BAD", "Common Equity", date.today(), 3, ""), db_path=temp_db)
    with pytest.raises(ValueError):
        build_package_version(package, db_path=temp_db)
    assert database.list_package_versions(package["package_id"], db_path=temp_db) == []


def test_original_files_remain_and_filename_collision_handling(temp_db: Path) -> None:
    package = create_package(PackageInput("COL", "Common Equity", date.today(), 3, ""), db_path=temp_db)
    package = database.update_package_company_metadata(package["package_id"], {"ticker": "COL", "company_name": "COL Inc.", "cik": "0001", "exchange": "NYSE", "sic": "", "industry_description": "", "fiscal_year_end": "", "sec_company_url": "", "resolution_status": "RESOLVED", "resolution_source": "test", "resolution_timestamp": "now"}, db_path=temp_db)
    store_uploaded_files(package, [UploadCandidate("same.pdf", b"%PDF a"), UploadCandidate("same.pdf", b"%PDF b")], source_type="other", authorization_confirmed=True, metadata_by_name={"same.pdf": {"final_category_code": "other"}}, db_path=temp_db)
    database.update_package_review_acknowledgement(package["package_id"], checklist_reviewed=True, reviewed_by="analyst", review_note="", missing_core_acknowledged=True, stale_documents_acknowledged=True, needs_review_acknowledged=True, db_path=temp_db)
    docs = database.list_documents_by_package(package["package_id"], db_path=temp_db)
    original_bytes = [Path(doc["local_path"]).read_bytes() for doc in docs if doc["collection_status"] == "DOWNLOADED"]
    version = build_package_version(database.get_package_by_package_id(package["package_id"], db_path=temp_db), db_path=temp_db)
    version_docs = database.list_package_version_documents(version["version_id"], db_path=temp_db)
    assert len({doc["package_filename"] for doc in version_docs}) == len(version_docs)
    assert [Path(doc["local_path"]).read_bytes() for doc in docs if doc["collection_status"] == "DOWNLOADED"] == original_bytes
    assert all(not Path(doc["relative_package_path"]).is_absolute() for doc in version_docs)


def test_integrity_detects_missing_changed_unexpected_size_hash(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    version = build_package_version(package, db_path=temp_db)
    root = Path(version["manifest_path"]).parents[1]
    docs = database.list_package_version_documents(version["version_id"], db_path=temp_db)
    assert verify_snapshot(root, docs)["overall_integrity_status"] == config.INTEGRITY_VERIFIED
    target = root / docs[0]["relative_package_path"]
    target.write_bytes(b"changed")
    report = verify_snapshot(root, docs)
    assert report["overall_integrity_status"] == config.INTEGRITY_FAILED
    assert report["hash_mismatches"] or report["size_mismatches"]
    target.unlink()
    assert verify_snapshot(root, docs)["missing_files"]
    (root / "unexpected.txt").write_text("extra", encoding="utf-8")
    assert "unexpected.txt" in verify_snapshot(root, docs)["unexpected_files"]


def test_checklist_snapshot_is_locked(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    version = build_package_version(package, db_path=temp_db)
    before = json.loads(version["checklist_snapshot_json"])
    set_override(package["package_id"], "latest_annual", config.CHECKLIST_STATUS_NOT_APPLICABLE, "changed", db_path=temp_db)
    after = json.loads(database.get_package_version(version["version_id"], db_path=temp_db)["checklist_snapshot_json"])
    assert before == after


def test_version_comparison_added_removed_renamed_recategorized_and_counts(temp_db: Path) -> None:
    package = prepared_package(temp_db)
    first = build_package_version(package, db_path=temp_db)
    doc = database.list_documents_by_package(package["package_id"], db_path=temp_db)[0]
    database.update_document_metadata(doc["document_id"], {"final_category_code": "other", "category": "Other"}, db_path=temp_db)
    store_uploaded_files(package, [UploadCandidate("new.pdf", b"%PDF new")], source_type="other", authorization_confirmed=True, metadata_by_name={"new.pdf": {"final_category_code": "other"}}, db_path=temp_db)
    second = build_package_version(database.get_package_by_package_id(package["package_id"], db_path=temp_db), db_path=temp_db)
    comparison = compare_versions(first["version_id"], second["version_id"], db_path=temp_db)
    assert comparison["documents_added"]
    assert comparison["documents_recategorized"]
    assert comparison["licensed_count_change"] >= 1


def test_phase4_streamlit_pages_load() -> None:
    for path in ["app/pages/3_Package_Review.py", "app/pages/5_Generated_Reports.py"]:
        app = AppTest.from_file(path, default_timeout=10)
        app.run()
        assert not list(app.exception)
