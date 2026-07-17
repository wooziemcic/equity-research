from __future__ import annotations

import hashlib
import json
import sqlite3
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app import config
from app.services.package_recipe_service import (
    activate_recipe,
    approve_recipe,
    assign_document,
    board_payload,
    classify_filename,
    clone_legacy_package,
    create_draft_version,
    create_package_from_active_recipe,
    database_audit_details,
    export_checklist_xlsx,
    export_package_snapshot,
    get_package_recipe_instance,
    import_package_snapshot,
    list_recipe_slots,
    mark_slot,
    recalculate_completion,
    suggest_document_assignments,
    update_assignment,
    update_draft_slot,
)
from app.services.package_service import PackageInput, create_package
from app.services.recipe_import_service import (
    COMMON_EQUITY_SHEETS,
    import_common_equity_recipe,
    inspect_common_equity_workbook,
)
from app.utils import database


SLOT_NAMES = [
    "BBG-DES", "BBG-DVD", "BBG-DDIS", "BBG-HDS", "Hoover Report",
    "Morningstar Report, and most recent Model", "BBG - ANR",
    "Sell Side Reports (GS)", "Credit Reports (Moodys)", "Initiated Coverage Report (GS)",
    "Independent Short Sale Report", "Industry Report", "Latest Earnings Release",
    "Available Supplemental", "Latest Earnings Call Transcript", "Investor Presentations",
    "Any company press releases since last earnings release",
    "Liquidity and Capital Resources Portion of last 10-Q or 10-K",
    "Description of Business and Risk", "Executive Compensation Information",
    "Most Recent 10-Q and 10-K", "CAST - Summary Chart", "DRSK - Default Risk",
    "CCM Historical Multiples valuation", "BBG - FA", "BBG - FA - Credit Ratios",
]


@pytest.fixture()
def synthetic_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "synthetic_cutler.xlsx"
    workbook = Workbook()
    template = workbook.active
    template.title = "Template"
    instructions = workbook.create_sheet("Instructions")
    mdt = workbook.create_sheet("MDT")
    convertible = workbook.create_sheet("cvt Checklist")
    convertible["A1"] = "Convertible-only source"
    for width, name in ((7, "Historical 7"), (8, "Historical 8"), (9, "Historical 9")):
        sheet = workbook.create_sheet(name)
        for column in range(1, width + 1):
            sheet.cell(1, column, f"value-{column}")

    requirements = ["Y", "N", "N/A", "-", ""]
    for index, name in enumerate(SLOT_NAMES, start=1):
        template_row = 6 + index
        template.cell(template_row, 5, name)
        template.cell(template_row, 7, requirements[index % len(requirements)])
        instruction_row = 17 + index
        instructions.cell(instruction_row, 1, index)
        instructions.cell(instruction_row, 2, name)
        instructions.cell(instruction_row, 3, requirements[(index + 1) % len(requirements)])
        instructions.cell(instruction_row, 4, "Synthetic source")
        instructions.cell(instruction_row, 6, f"Synthetic instruction for {name}")

    mdt["C1"] = "Comprehensive Equity Research Package"
    row = 7
    order = 1
    for index, name in enumerate(SLOT_NAMES):
        if name == "CAST - Summary Chart":
            order = 23
        mdt.cell(row, 3, order)
        mdt.cell(row, 4, name)
        mdt.cell(row, 6, requirements[index % len(requirements)])
        mdt.cell(row, 7, "Synthetic source")
        row += 1
        if order < 21:
            order += 1
        elif order >= 23:
            order += 1
    mdt.insert_rows(17)
    mdt.cell(17, 4, "Sell Side Downgrade")
    mdt.cell(17, 6, "N/A")
    mdt.insert_rows(23)
    mdt.cell(23, 4, "Latest Earnings Call Audio")
    mdt.cell(23, 6, "Y")
    workbook.save(path)
    return path


@pytest.fixture()
def active_recipe_db(tmp_path: Path, synthetic_workbook: Path) -> tuple[Path, dict]:
    db_path = tmp_path / "phase6a.db"
    imported = import_common_equity_recipe(synthetic_workbook, imported_by="unit-admin", db_path=db_path)
    approve_recipe(imported["recipe_id"], approver="unit-admin", db_path=db_path)
    activate_recipe(imported["recipe_id"], actor="unit-admin", db_path=db_path)
    return db_path, imported


def _create_recipe_package(db_path: Path, ticker: str = "MDT") -> dict:
    return create_package_from_active_recipe(
        {"ticker": ticker, "company_name": "Synthetic Company", "cik": "0001613103", "exchange": "NYSE", "resolution_status": "RESOLVED"},
        research_cutoff=date.today(), compilation_date=date.today(), compiled_by="Unit Analyst", created_by="Unit Analyst", db_path=db_path,
    )


def _document(db_path: Path, package: dict, filename: str, suffix: str = "A") -> dict:
    return database.create_document_record(
        {
            "document_id": f"DOC-{suffix}", "package_id": package["package_id"], "ticker": package["ticker"],
            "category": "Licensed Research", "document_type": "Research", "title": Path(filename).stem,
            "source_name": "Synthetic", "source_url": "", "publication_date": "2026-07-01",
            "local_filename": filename, "mime_type": "application/pdf", "file_size_bytes": 10,
            "sha256_hash": hashlib.sha256(filename.encode()).hexdigest(), "collection_method": "UPLOAD",
            "collection_status": "DOWNLOADED", "is_public": False, "original_filename": filename,
        },
        db_path=db_path,
    )


def test_workbook_import_preserves_provenance_ambiguity_and_order(synthetic_workbook: Path) -> None:
    report = inspect_common_equity_workbook(synthetic_workbook)
    assert set(COMMON_EQUITY_SHEETS).issubset(report["available_sheets"])
    assert "cvt Checklist" not in report["selected_sheets"]
    assert report["workbook_sha256"] == hashlib.sha256(synthetic_workbook.read_bytes()).hexdigest()
    assert report["raw_rows"]["MDT"][0]["coordinates"]
    assert report["missing_order_numbers"] == [22]
    assert not any(slot["order_number"] == 22 for slot in report["slots"])
    assert report["unnumbered_supplemental_slots"] == 2
    assert [(slot["display_name"], slot["suborder"]) for slot in report["slots"] if slot["suborder"]] == [
        ("Sell-Side Downgrade", 1), ("Latest Earnings Call Audio", 1)
    ]
    raw_requirements = {slot["raw_import"]["raw_requirement"] for slot in report["slots"]}
    assert {"Y", "N", "N/A", "-", ""}.issubset(raw_requirements)
    assert any(warning["code"] == "AMBIGUOUS_REQUIREMENT" for warning in report["warnings"])
    assert report["normalized_slot_count"] == 28


def test_import_detects_duplicate_labels(tmp_path: Path, synthetic_workbook: Path) -> None:
    workbook = load_workbook(synthetic_workbook)
    workbook["MDT"]["D7"] = "BBG-DVD"
    duplicate_path = tmp_path / "duplicate.xlsx"
    workbook.save(duplicate_path)
    report = inspect_common_equity_workbook(duplicate_path)
    assert any(warning["code"] == "DUPLICATE_LABEL" for warning in report["warnings"])


def test_recipe_approval_is_immutable_and_new_edits_require_version(tmp_path: Path, synthetic_workbook: Path) -> None:
    db_path = tmp_path / "recipes.db"
    imported = import_common_equity_recipe(synthetic_workbook, imported_by="admin", db_path=db_path)
    first_slot = list_recipe_slots(imported["recipe_id"], db_path=db_path)[0]
    update_draft_slot(first_slot["slot_id"], {"display_name": "Reviewed BBG-DES"}, actor="admin", db_path=db_path)
    approved = approve_recipe(imported["recipe_id"], approver="admin", db_path=db_path)
    assert approved["status"] == "APPROVED"
    with pytest.raises(ValueError, match="immutable"):
        update_draft_slot(first_slot["slot_id"], {"display_name": "Forbidden"}, actor="admin", db_path=db_path)
    draft = create_draft_version(imported["recipe_id"], created_by="admin", db_path=db_path)
    assert draft["version"] == imported["version"] + 1
    assert draft["status"] == "NEEDS_REVIEW"


def test_migration_is_additive_idempotent_and_test_database_isolated(tmp_path: Path) -> None:
    runtime_path = Path(config.DATABASE_PATH)
    runtime_state = (runtime_path.exists(), runtime_path.stat().st_mtime_ns if runtime_path.exists() else None)
    db_path = tmp_path / "isolated.db"
    database.initialize_database(db_path)
    legacy = create_package(PackageInput("QXO", "Common Equity", date.today(), 3), db_path=db_path)
    before = database.get_package_by_package_id(legacy["package_id"], db_path=db_path)
    database.initialize_database(db_path)
    after = database.get_package_by_package_id(legacy["package_id"], db_path=db_path)
    assert before == after
    details = database_audit_details(db_path=db_path)
    assert details == {"environment": "TEST", "storage": "Temporary", "package_count": 1, "most_recent_package_id": legacy["package_id"], "schema_version": "6C.0"}
    assert runtime_state == (runtime_path.exists(), runtime_path.stat().st_mtime_ns if runtime_path.exists() else None)


def test_development_migration_creates_one_safe_backup(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    db_path = tmp_path / "development.db"
    backup_dir = tmp_path / "migration_backups"
    with sqlite3.connect(db_path) as connection:
        connection.execute("CREATE TABLE sentinel(value TEXT)")
        connection.execute("INSERT INTO sentinel VALUES ('historical-data')")
    monkeypatch.setattr(database, "DATABASE_PATH", db_path)
    monkeypatch.setattr(config, "DATABASE_ENVIRONMENT", "DEVELOPMENT")
    monkeypatch.setattr(config, "MIGRATION_BACKUP_DIR", backup_dir)
    database.initialize_database(db_path)
    backups = list(backup_dir.glob("cutler_research_pre_phase6a_*.db"))
    assert len(backups) == 1
    with sqlite3.connect(backups[0]) as connection:
        assert connection.execute("SELECT value FROM sentinel").fetchone()[0] == "historical-data"
        assert not connection.execute("SELECT 1 FROM sqlite_master WHERE name='package_recipes'").fetchone()
    database.initialize_database(db_path)
    assert list(backup_dir.glob("cutler_research_pre_phase6a_*.db")) == backups


def test_package_snapshot_and_recipe_order_are_immutable(active_recipe_db: tuple[Path, dict]) -> None:
    db_path, imported = active_recipe_db
    package = _create_recipe_package(db_path)
    instance = get_package_recipe_instance(package["package_id"], db_path=db_path)
    original_snapshot = instance["recipe_snapshot_json"]
    assert instance["recipe_version"] == imported["version"]
    slots = board_payload(package["package_id"], db_path=db_path)["slots"]
    assert [slot["order_number"] for slot in slots if slot["suborder"] == 0][-6:] == [21, 23, 24, 25, 26, 27]
    draft = create_draft_version(imported["recipe_id"], created_by="next-admin", db_path=db_path)
    assert draft["version"] == 2
    assert get_package_recipe_instance(package["package_id"], db_path=db_path)["recipe_snapshot_json"] == original_snapshot


@pytest.mark.parametrize(
    ("filename", "expected", "review"),
    [
        ("MDT GS 6.4.26.pdf", "Sell-Side Reports", True),
        ("MDT Moody's 10.11.24.pdf", "Credit Reports", False),
        ("MDT Evercore 7.6.26 Industry.pdf", "Industry Report", False),
        ("MDT GS 2.12.20 Initiation.pdf", "Initiated Coverage Report", False),
        ("MDT Earnings-Commentary-FY26Q4 6.3.26.pdf", "Latest Earnings Call Transcript", True),
        ("MDT Earnings Presentation.pdf", "Available Supplemental or Earnings Presentation", False),
        ("MDT Credit Ratios 7.9.26.pdf", "BBG-FA Credit Ratios", False),
        ("MDT DRSK 7.9.26.pdf", "DRSK Default Risk", False),
        ("MDT FA 7.9.26.pdf", "BBG-FA", False),
        ("MDT ANR 7.9.26.xlsm", "BBG-ANR", False),
        ("MDT Morningstar.pdf", "Morningstar Report and Most Recent Model", False),
        ("MDT unknown.pdf", None, True),
    ],
)
def test_deterministic_filename_classification(filename: str, expected: str | None, review: bool) -> None:
    names = {
        "Sell-Side Reports", "Credit Reports", "Industry Report", "Initiated Coverage Report",
        "Latest Earnings Call Transcript", "Available Supplemental or Earnings Presentation",
        "BBG-FA Credit Ratios", "DRSK Default Risk", "BBG-FA", "BBG-ANR",
        "Morningstar Report and Most Recent Model",
    }
    slots = [{"display_name": name, "slot_id": f"slot-{index}"} for index, name in enumerate(names)]
    result = classify_filename(filename, slots)
    assert result["suggested_slot_name"] == expected
    assert result["requires_review"] is review


def test_assignments_caps_completion_and_rejected_files(active_recipe_db: tuple[Path, dict]) -> None:
    db_path, _ = active_recipe_db
    package = _create_recipe_package(db_path)
    payload = board_payload(package["package_id"], db_path=db_path)
    slot = next(item for item in payload["slots"] if item["display_name_snapshot"] == "Credit Reports")
    first = _document(db_path, package, "MDT Moody's 10.11.24.pdf", "ONE")
    second = _document(db_path, package, "MDT S&P Credit.pdf", "TWO")
    assignment = assign_document(slot["package_slot_instance_id"], first["document_id"], actor="analyst", db_path=db_path)
    assert assignment["assignment_status"] == "APPROVED"
    industry = next(item for item in payload["slots"] if item["display_name_snapshot"] == "Industry Report")
    with pytest.raises(ValueError, match="generated derivative"):
        assign_document(industry["package_slot_instance_id"], first["document_id"], actor="analyst", db_path=db_path)
    for index in range(2, 6):
        extra = _document(db_path, package, f"MDT Moody Credit {index}.pdf", f"EX{index}")
        assign_document(slot["package_slot_instance_id"], extra["document_id"], actor="analyst", db_path=db_path)
    with pytest.raises(ValueError, match="cap"):
        assign_document(slot["package_slot_instance_id"], second["document_id"], actor="analyst", db_path=db_path)
    override = assign_document(slot["package_slot_instance_id"], second["document_id"], actor="admin", override_cap=True, override_reason="Approved expanded credit review.", db_path=db_path)
    update_assignment(override["assignment_id"], "reject", actor="analyst", db_path=db_path)
    refreshed = recalculate_completion(package["package_id"], db_path=db_path)
    assert refreshed["recommended_complete"] >= 1


def test_not_applicable_and_optional_slots_do_not_reduce_required_readiness(active_recipe_db: tuple[Path, dict]) -> None:
    db_path, _ = active_recipe_db
    package = _create_recipe_package(db_path)
    payload = board_payload(package["package_id"], db_path=db_path)
    conditional = next(slot for slot in payload["slots"] if slot["requirement_snapshot"] == "CONDITIONAL")
    mark_slot(conditional["package_slot_instance_id"], "NOT_APPLICABLE", reason="Not applicable to this domestic issuer.", actor="analyst", db_path=db_path)
    summary = recalculate_completion(package["package_id"], db_path=db_path)
    assert summary["overall_total"] == len(payload["slots"]) - 1
    assert summary["required_total"] == 5


def test_checklist_export_and_snapshot_are_database_derived_and_secret_free(active_recipe_db: tuple[Path, dict], synthetic_workbook: Path) -> None:
    db_path, _ = active_recipe_db
    source_hash = hashlib.sha256(synthetic_workbook.read_bytes()).hexdigest()
    package = _create_recipe_package(db_path)
    slots = board_payload(package["package_id"], db_path=db_path)["slots"]
    credit = next(slot for slot in slots if slot["display_name_snapshot"] == "Credit Reports")
    document = _document(db_path, package, "MDT Moody's 10.11.24.pdf")
    assign_document(credit["package_slot_instance_id"], document["document_id"], actor="analyst", db_path=db_path)
    content = export_checklist_xlsx(package["package_id"], actor="analyst", db_path=db_path)
    export_path = synthetic_workbook.parent / "export.xlsx"
    export_path.write_bytes(content)
    sheet = load_workbook(export_path, data_only=True)["Package Checklist"]
    values = [sheet.cell(row, 1).value for row in range(9, sheet.max_row + 1)]
    assert "21" in values and "23" in values and "22" not in values
    assert any(sheet.cell(row, 6).value == "MDT Moody's 10.11.24" for row in range(9, sheet.max_row + 1))
    snapshot = export_package_snapshot(package["package_id"], db_path=db_path)
    assert b"api_key" not in snapshot.lower() and b"authorization" not in snapshot.lower()
    assert hashlib.sha256(synthetic_workbook.read_bytes()).hexdigest() == source_hash


def test_legacy_clone_and_snapshot_import_create_distinct_drafts(active_recipe_db: tuple[Path, dict]) -> None:
    db_path, _ = active_recipe_db
    legacy = create_package(PackageInput("LEG", "Common Equity", date.today(), 3), db_path=db_path)
    original = database.get_package_by_package_id(legacy["package_id"], db_path=db_path)
    cloned = clone_legacy_package(legacy["package_id"], created_by="analyst", db_path=db_path)
    assert cloned["package_id"] != legacy["package_id"]
    assert cloned["source_legacy_package_id"] == legacy["package_id"]
    assert database.get_package_by_package_id(legacy["package_id"], db_path=db_path) == original
    snapshot = export_package_snapshot(cloned["package_id"], db_path=db_path)
    imported = import_package_snapshot(snapshot, imported_by="admin", db_path=db_path)
    assert imported["package_id"] not in {legacy["package_id"], cloned["package_id"]}
    assert get_package_recipe_instance(imported["package_id"], db_path=db_path)


def test_board_contract_has_sections_mobile_css_and_targeted_search(active_recipe_db: tuple[Path, dict]) -> None:
    db_path, _ = active_recipe_db
    package = _create_recipe_package(db_path)
    payload = board_payload(package["package_id"], db_path=db_path)
    sections = list(dict.fromkeys(slot["section_snapshot"] for slot in payload["slots"]))
    assert sections == [
        "Company Snapshot", "Licensed and Third-Party Research", "Earnings and Company Materials",
        "SEC Filings and Extracted Sections", "Internal Cutler Analysis",
    ]
    page = (config.PROJECT_ROOT / "app" / "pages" / "8_Package_Assembly.py").read_text(encoding="utf-8")
    css = config.STYLE_PATH.read_text(encoding="utf-8")
    assert "Comprehensive Equity Research Package" in page
    assert 'st.button("Find Automatically"' in page
    assert ".assembly-mobile" in css and "@media (max-width: 760px)" in css
